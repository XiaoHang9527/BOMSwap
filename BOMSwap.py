import argparse
import logging
import pandas as pd
from pathlib import Path
from pathlib import Path
import os
import time
import sys
import json
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import tkinter as tk
from tkinter import filedialog, ttk, messagebox, StringVar
import tkinter.messagebox
from threading import Thread
import traceback  # 增加traceback模块用于详细错误信息

# 添加更新功能所需的库
import requests
import tempfile
import shutil
import zipfile
import subprocess
import platform
from packaging import version as pkg_version

# 定义版本信息和更新相关常量
APP_NAME = "BOM替代料工具"
APP_VERSION = "2.5"
GITHUB_REPO = "XiaoHang9527/BOMSwap"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
UPDATE_CHECK_INTERVAL = 7  # 天

# 定义下载重试次数和超时时间
DOWNLOAD_MAX_RETRIES = 3  # 最大重试次数
DOWNLOAD_TIMEOUT = 30     # 下载超时时间(秒)
DOWNLOAD_CHUNK_SIZE = 8192  # 下载块大小

# 配置文件路径
# 在程序目录下创建配置文件
def get_program_dir():
    """获取程序目录"""
    try:
        # 首先尝试获取exe文件所在目录（打包环境）
        if getattr(sys, 'frozen', False):
            exe_dir = os.path.dirname(sys.executable)
            print(f"使用exe所在目录: {exe_dir}")
            logging.info(f"使用exe所在目录: {exe_dir}")
            return exe_dir

        # 如果不是frozen环境，尝试使用resource_path获取程序目录
        base_path = os.path.dirname(resource_path(""))
        print(f"使用打包环境程序目录: {base_path}")
        logging.info(f"使用打包环境程序目录: {base_path}")
        return base_path
    except Exception as e:
        # 如果上述方法都失败，使用__file__获取程序目录（开发环境）
        try:
            program_dir = os.path.dirname(os.path.abspath(__file__))
            print(f"使用开发环境程序目录: {program_dir}")
            logging.info(f"使用开发环境程序目录: {program_dir}")
            return program_dir
        except Exception as e2:
            # 如果所有方法都失败，使用当前工作目录
            current_dir = os.getcwd()
            print(f"所有方法获取程序目录失败，使用当前工作目录: {current_dir}, 错误: {e}, {e2}")
            logging.error(f"所有方法获取程序目录失败，使用当前工作目录: {current_dir}, 错误: {e}, {e2}")
            return current_dir

# 程序目录下的配置文件
CONFIG_FILE = os.path.join(get_program_dir(), "config.json")

# 性能优化：缓存一些常用的配置和计算结果
_config_cache = None
_default_font = None
_config_file_path = None  # 保存成功加载的配置文件路径

# 定义全局颜色变量
header_bg_color = "0078D4"  # 微软蓝
default_highlight_color = "FFFFC0"  # 浅黄色，用于替代料

# 删除不再需要的ensure_config_dir函数

def check_directory_writable(directory):
    """
    检查目录是否可写，通过尝试创建一个临时文件

    Args:
        directory: 要检查的目录路径

    Returns:
        bool: 目录是否可写
    """
    if not os.path.exists(directory):
        try:
            os.makedirs(directory)
            print(f"创建目录成功: {directory}")
            logging.info(f"创建目录成功: {directory}")
        except Exception as e:
            print(f"创建目录失败: {e}")
            logging.error(f"创建目录失败: {e}")
            return False

    # 检查目录是否可写
    if not os.access(directory, os.W_OK):
        print(f"警告: 目录不可写: {directory}")
        logging.warning(f"目录不可写: {directory}")
        return False

    # 尝试创建临时文件
    test_file = os.path.join(directory, "write_test.tmp")
    try:
        with open(test_file, 'w') as f:
            f.write("测试写入权限")
        os.remove(test_file)
        print(f"目录可写: {directory}")
        logging.info(f"目录可写: {directory}")
        return True
    except Exception as e:
        print(f"目录写入测试失败: {directory}, 错误: {e}")
        logging.error(f"目录写入测试失败: {directory}, 错误: {e}")
        return False

def save_config(config):
    """保存配置到文件"""
    global _config_cache, _config_file_path, CONFIG_FILE

    # 更新缓存
    _config_cache = config

    # 获取程序目录
    program_dir = get_program_dir()
    print(f"尝试保存配置到程序目录: {program_dir}")
    logging.info(f"尝试保存配置到程序目录: {program_dir}")

    # 尝试的保存路径列表
    save_paths = []

    # 1. 首先尝试程序目录
    save_paths.append((program_dir, "程序目录"))

    # 2. 然后尝试当前工作目录（如果与程序目录不同）
    current_dir = os.getcwd()
    if program_dir != current_dir:
        save_paths.append((current_dir, "当前工作目录"))

    # 3. 最后尝试用户文档目录
    try:
        user_docs = os.path.join(os.path.expanduser("~"), "Documents")
        if os.path.exists(user_docs) and program_dir != user_docs and current_dir != user_docs:
            save_paths.append((user_docs, "用户文档目录"))
    except Exception as e:
        print(f"获取用户文档目录失败: {e}")
        logging.error(f"获取用户文档目录失败: {e}")

    # 4. 尝试用户主目录
    try:
        user_home = os.path.expanduser("~")
        if program_dir != user_home and current_dir != user_home:
            save_paths.append((user_home, "用户主目录"))
    except Exception as e:
        print(f"获取用户主目录失败: {e}")
        logging.error(f"获取用户主目录失败: {e}")

    # 依次尝试每个路径
    for save_dir, dir_desc in save_paths:
        try:
            # 检查目录是否可写
            if check_directory_writable(save_dir):
                config_path = os.path.join(save_dir, 'config.json')

                # 尝试保存配置
                try:
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=4)

                    print(f"配置已成功保存到{dir_desc}: {config_path}")
                    logging.info(f"配置已成功保存到{dir_desc}: {config_path}")

                    # 更新成功加载的配置文件路径
                    _config_file_path = config_path

                    # 如果不是保存到程序目录，且程序目录是CONFIG_FILE的目录，更新CONFIG_FILE
                    if save_dir != program_dir and os.path.dirname(CONFIG_FILE) == program_dir:
                        CONFIG_FILE = config_path
                        print(f"已更新CONFIG_FILE路径为: {CONFIG_FILE}")
                        logging.info(f"已更新CONFIG_FILE路径为: {CONFIG_FILE}")

                    return True
                except Exception as e:
                    print(f"保存配置到{dir_desc}失败: {e}")
                    logging.error(f"保存配置到{dir_desc}失败: {e}")
                    # 继续尝试下一个路径
            else:
                print(f"{dir_desc}不可写: {save_dir}")
                logging.warning(f"{dir_desc}不可写: {save_dir}")
        except Exception as e:
            print(f"检查{dir_desc}可写性时出错: {e}")
            logging.error(f"检查{dir_desc}可写性时出错: {e}")

    # 所有路径都尝试失败
    print("所有尝试的路径都无法保存配置文件")
    logging.error("所有尝试的路径都无法保存配置文件")

    # 最后尝试直接保存到CONFIG_FILE指定的路径
    try:
        # 确保目录存在
        config_dir = os.path.dirname(CONFIG_FILE)
        if not os.path.exists(config_dir):
            os.makedirs(config_dir)

        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)

        print(f"配置已成功保存到CONFIG_FILE路径: {CONFIG_FILE}")
        logging.info(f"配置已成功保存到CONFIG_FILE路径: {CONFIG_FILE}")

        # 更新成功加载的配置文件路径
        _config_file_path = CONFIG_FILE
        return True
    except Exception as e:
        print(f"保存配置到CONFIG_FILE路径失败: {e}")
        logging.error(f"保存配置到CONFIG_FILE路径失败: {e}")
        return False

def load_default_config(use_builtin_defaults=False):
    """
    从默认配置文件加载默认配置

    Args:
        use_builtin_defaults: 是否使用内置默认配置，True表示强制使用内置默认配置，不读取外部文件
    """
    # 如果要求使用内置默认配置，直接返回内置默认值
    if use_builtin_defaults:
        print("使用内置默认配置")
        logging.info("使用内置默认配置")
        return get_builtin_default_config()

    # 只使用程序目录下的config.json
    possible_paths = []

    # 程序目录下的config.json
    program_dir = get_program_dir()
    program_config = os.path.join(program_dir, 'config.json')
    possible_paths.append(program_config)
    print(f"使用程序目录配置路径: {program_config}")
    logging.info(f"使用程序目录配置路径: {program_config}")

    # 如果程序目录与当前工作目录不同，也尝试从当前工作目录加载
    current_dir = os.getcwd()
    if program_dir != current_dir:
        current_config = os.path.join(current_dir, 'config.json')
        possible_paths.append(current_config)
        print(f"备用：当前工作目录配置路径: {current_config}")
        logging.info(f"备用：当前工作目录配置路径: {current_config}")

    # 打印当前工作目录和程序目录，帮助调试
    print(f"当前工作目录: {os.getcwd()}")
    print(f"程序目录: {os.path.dirname(os.path.abspath(__file__))}")

    # 尝试从每个可能的路径加载配置
    for config_path in possible_paths:
        print(f"尝试加载配置文件: {config_path}")
        logging.info(f"尝试加载配置文件: {config_path}")

        if os.path.exists(config_path):
            print(f"配置文件存在: {config_path}")
            logging.info(f"配置文件存在: {config_path}")

            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    file_content = f.read()
                    print(f"配置文件内容: {file_content}")

                    # 重新打开文件进行JSON解析
                    f.seek(0)
                    default_settings = json.load(f)

                # 打印加载的配置内容，用于调试
                print(f"从文件加载的配置内容: {default_settings}")

                # 检查配置文件中是否包含必要的键
                if 'bom_header_mapping' not in default_settings or 'sub_header_mapping' not in default_settings:
                    print(f"配置文件 {config_path} 缺少必要的键，跳过")
                    continue

                # 添加其他必要的默认配置项
                default_config = {
                    'last_bom_dir': '',
                    'default_sub_path': '',
                    'bom_header_mapping': default_settings['bom_header_mapping'],
                    'sub_header_mapping': default_settings['sub_header_mapping'],
                    'highlight_color': default_settings.get('highlight_color', default_highlight_color),
                    'last_update_check': 0,  # 上次检查更新的时间戳
                    'last_used_header_mapping': {}  # 上次使用的表头映射
                }

                print(f"从配置文件加载配置成功: {config_path}")
                print(f"加载的BOM表头映射: {default_config['bom_header_mapping']}")
                print(f"加载的替代料表表头映射: {default_config['sub_header_mapping']}")
                logging.info(f"从配置文件加载配置成功: {config_path}")

                # 将成功加载的配置文件路径保存到全局变量，方便后续保存配置
                global _config_file_path
                _config_file_path = config_path

                return default_config
            except Exception as e:
                print(f"加载配置文件失败: {config_path}, 错误: {e}")
                logging.error(f"加载配置文件失败: {config_path}, 错误: {e}")
                # 继续尝试下一个路径

    # 如果所有路径都加载失败，使用内置默认配置
    print("所有配置文件加载失败，使用内置默认配置")
    logging.info("所有配置文件加载失败，使用内置默认配置")
    return get_builtin_default_config()

def get_builtin_default_config():
    """获取内置默认配置"""
    return {
        'last_bom_dir': '',
        'default_sub_path': '',  # 替代料关系表的默认路径
        'bom_header_mapping': {
            'item': 'Item',
            'pn': 'PN',
            'part': 'Part',
            'reference': 'Reference',
            'quantity': 'Quantity',
            'description': 'Description',
            'mfr_pn': 'ManufacturerPN',
            'manufacturer': 'Manufacturer'
        },
        'sub_header_mapping': {
            'pn': 'PN',
            'part': 'Part',
            'description': 'Description',
            'mfr_pn': 'ManufacturerPN',
            'manufacturer': 'Manufacturer',
            'attribute': 'attribute'
        },
        'highlight_color': default_highlight_color,
        'last_update_check': 0,
        'last_used_header_mapping': {}
    }

def load_config():
    """加载配置文件，如果不存在则创建默认配置"""
    global _config_cache, _config_file_path, CONFIG_FILE

    # 如果缓存存在且有效，直接返回缓存
    if _config_cache is not None:
        print("使用缓存的配置")
        return _config_cache

    # 加载默认配置
    default_config = load_default_config()

    # 打印默认配置中的表头映射，用于调试
    print("默认配置中的BOM表头映射:", default_config['bom_header_mapping'])
    print("默认配置中的替代料表表头映射:", default_config['sub_header_mapping'])

    # 尝试加载的配置文件路径列表
    config_paths = []

    # 1. 如果有成功加载的配置文件路径，优先使用该路径
    if _config_file_path and os.path.exists(_config_file_path):
        config_paths.append((_config_file_path, "已知配置文件路径"))

    # 2. 然后尝试CONFIG_FILE路径
    if os.path.exists(CONFIG_FILE):
        if not _config_file_path or CONFIG_FILE != _config_file_path:
            config_paths.append((CONFIG_FILE, "CONFIG_FILE路径"))

    # 3. 尝试程序目录下的config.json
    program_dir = get_program_dir()
    program_config = os.path.join(program_dir, 'config.json')
    if os.path.exists(program_config) and program_config not in [p[0] for p in config_paths]:
        config_paths.append((program_config, "程序目录"))

    # 4. 尝试当前工作目录下的config.json
    current_dir = os.getcwd()
    current_config = os.path.join(current_dir, 'config.json')
    if os.path.exists(current_config) and current_config not in [p[0] for p in config_paths]:
        config_paths.append((current_config, "当前工作目录"))

    # 5. 尝试用户文档目录下的config.json
    try:
        user_docs = os.path.join(os.path.expanduser("~"), "Documents")
        user_docs_config = os.path.join(user_docs, 'config.json')
        if os.path.exists(user_docs_config) and user_docs_config not in [p[0] for p in config_paths]:
            config_paths.append((user_docs_config, "用户文档目录"))
    except Exception as e:
        print(f"获取用户文档目录失败: {e}")
        logging.error(f"获取用户文档目录失败: {e}")

    # 6. 尝试用户主目录下的config.json
    try:
        user_home = os.path.expanduser("~")
        user_home_config = os.path.join(user_home, 'config.json')
        if os.path.exists(user_home_config) and user_home_config not in [p[0] for p in config_paths]:
            config_paths.append((user_home_config, "用户主目录"))
    except Exception as e:
        print(f"获取用户主目录失败: {e}")
        logging.error(f"获取用户主目录失败: {e}")

    # 依次尝试每个配置文件路径
    for config_path, path_desc in config_paths:
        print(f"尝试从{path_desc}加载配置: {config_path}")
        logging.info(f"尝试从{path_desc}加载配置: {config_path}")

        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                file_content = f.read()
                print(f"配置文件内容: {file_content[:100]}..." if len(file_content) > 100 else file_content)

                # 重新打开文件进行JSON解析
                f.seek(0)
                user_config = json.load(f)

            # 确保所有默认键都存在
            for key, value in default_config.items():
                if key not in user_config:
                    user_config[key] = value
                # 对于嵌套字典，确保所有默认键都存在
                elif isinstance(value, dict) and key in user_config:
                    for sub_key, sub_value in value.items():
                        if sub_key not in user_config[key]:
                            user_config[key][sub_key] = sub_value

            # 打印加载的配置内容
            print(f"从{path_desc}加载配置成功: {config_path}")
            print(f"加载的BOM表头映射: {user_config['bom_header_mapping']}")
            print(f"加载的替代料表表头映射: {user_config['sub_header_mapping']}")
            logging.info(f"从{path_desc}加载配置成功: {config_path}")

            # 更新成功加载的配置文件路径
            _config_file_path = config_path

            # 如果加载的不是CONFIG_FILE路径，更新CONFIG_FILE
            if config_path != CONFIG_FILE:
                CONFIG_FILE = config_path
                print(f"已更新CONFIG_FILE路径为: {CONFIG_FILE}")
                logging.info(f"已更新CONFIG_FILE路径为: {CONFIG_FILE}")

            _config_cache = user_config
            return user_config
        except Exception as e:
            print(f"从{path_desc}加载配置失败: {e}")
            logging.error(f"从{path_desc}加载配置失败: {e}")
            # 继续尝试下一个路径

    # 所有路径都加载失败，创建默认配置
    print("所有配置文件加载失败，创建默认配置")
    logging.info("所有配置文件加载失败，创建默认配置")

    # 保存默认配置
    save_success = save_config(default_config)
    if save_success:
        print("成功创建默认配置文件")
        logging.info("成功创建默认配置文件")
    else:
        print("创建默认配置文件失败，将使用内存中的默认配置")
        logging.warning("创建默认配置文件失败，将使用内存中的默认配置")

    _config_cache = default_config
    return default_config

def resource_path(relative_path):
    """获取资源的绝对路径，适用于开发环境和PyInstaller打包后的环境"""
    try:
        # PyInstaller创建临时文件夹，将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
        print(f"使用PyInstaller打包环境临时路径: {base_path}")
        logging.info(f"使用PyInstaller打包环境临时路径: {base_path}")

        # 对于配置文件，我们需要使用应用程序的实际安装目录，而不是临时目录
        if relative_path == "config.json":
            # 获取exe文件所在目录
            exe_dir = os.path.dirname(sys.executable)
            print(f"检测到配置文件请求，使用exe所在目录: {exe_dir}")
            logging.info(f"检测到配置文件请求，使用exe所在目录: {exe_dir}")
            return os.path.join(exe_dir, relative_path)
    except Exception as e:
        # 如果不是打包环境，使用当前路径
        base_path = os.path.abspath(".")
        print(f"使用开发环境路径: {base_path}, 原因: {str(e)}")
        logging.info(f"使用开发环境路径: {base_path}")

    full_path = os.path.join(base_path, relative_path)

    # 如果是空路径，只返回基础路径
    if not relative_path:
        return base_path

    # 检查路径是否存在
    if relative_path and not os.path.exists(full_path) and relative_path != "config.json":
        print(f"警告: 资源路径不存在: {full_path}")
        logging.warning(f"资源路径不存在: {full_path}")

    return full_path

def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        stream=None
    )

def parse_args():
    parser = argparse.ArgumentParser(description='BOM替代料工具')
    parser.add_argument('-i', '--input', required=True, help='输入BOM文件路径')
    parser.add_argument('-s', '--sub', required=True, help='替代料表文件路径')
    parser.add_argument('-o', '--output', help='输出文件路径')
    return parser.parse_args()

import tkinter as tk
from tkinter import filedialog, ttk
from threading import Thread
import tkinter.messagebox

def create_gui():
    global root, status_var, progress, bom_var, sub_var, update_manager
    root = tk.Tk()
    root.withdraw()
    try:
        # 获取图标资源路径
        icon_path = resource_path('DSC01.ico')

        # 尝试加载图标
        if os.path.exists(icon_path):
            root.iconbitmap(icon_path)
        else:
            # 如果资源路径找不到，尝试其他可能的位置
            alternative_paths = [
                os.path.join(os.path.dirname(os.path.abspath(__file__)), 'DSC01.ico'),
                r'D:\AI_Code\BOM替代料工具\DSC01.ico',
                './DSC01.ico'
            ]

            for path in alternative_paths:
                if os.path.exists(path):
                    root.iconbitmap(path)
                    break
    except Exception as e:
        print(f"图标加载错误: {e}")
        # 图标加载失败时继续运行程序
        pass
    root.title(f'BOM替代料工具 v{APP_VERSION} | 小航  2025.5.12')

    # 创建更新管理器
    global update_manager
    update_manager = UpdateManager(root)

    # 设置UI主题和样式
    style = ttk.Style()
    style.theme_use('clam')  # 使用clam主题作为基础

    # macOS风格颜色方案
    mac_bg = "#F5F5F7"  # 背景色 - 浅灰色
    mac_accent = "#0066CC"  # 强调色 - Apple蓝
    mac_secondary = "#E8E8ED"  # 次要背景色
    mac_text = "#1D1D1F"  # 主文本色
    mac_subtle_text = "#86868B"  # 次要文本色
    mac_border = "#D2D2D7"  # 边框色
    mac_button_bg = "#FFFFFF"  # 按钮背景色
    mac_selection = "#E8F0FE"  # 选中项背景色

    # 配置基本样式
    default_font_family = "微软雅黑"  # Windows平台使用微软雅黑

    # 设置窗口背景色
    root.configure(background=mac_bg)

    # 配置样式
    style.configure('.', background=mac_bg)
    style.configure('TFrame', background=mac_bg)
    style.configure('TLabel', background=mac_bg, foreground=mac_text, font=(default_font_family, 10))
    style.configure('TLabelframe', background=mac_bg, foreground=mac_text, bordercolor=mac_border)
    style.configure('TLabelframe.Label', background=mac_bg, foreground=mac_text, font=(default_font_family, 10, 'bold'))

    # 配置Entry样式
    style.configure('TEntry', background=mac_button_bg, fieldbackground=mac_button_bg, foreground=mac_text,
                    bordercolor=mac_border, lightcolor=mac_border, darkcolor=mac_border,
                    borderwidth=1, arrowsize=12)
    style.map('TEntry',
              bordercolor=[('focus', mac_accent)],
              lightcolor=[('focus', mac_accent)],
              darkcolor=[('focus', mac_accent)])

    # 配置按钮样式 - 标准按钮和强调按钮
    style.configure('TButton',
                    background=mac_button_bg,
                    foreground=mac_text,
                    bordercolor=mac_border,
                    lightcolor=mac_button_bg,
                    darkcolor=mac_button_bg,
                    borderwidth=1,
                    padding=(16, 8),
                    relief='flat',
                    font=(default_font_family, 10))
    style.map('TButton',
              background=[('active', mac_button_bg), ('pressed', mac_selection)],
              foreground=[('active', mac_text), ('pressed', mac_text)],
              bordercolor=[('active', mac_accent), ('pressed', mac_accent)])

    # 配置强调按钮样式
    style.configure('Accent.TButton',
                    background=mac_accent,
                    foreground='white',
                    bordercolor=mac_accent,
                    lightcolor=mac_accent,
                    darkcolor=mac_accent,
                    borderwidth=0,
                    relief='flat')
    style.map('Accent.TButton',
              background=[('active', '#0055B0'), ('pressed', '#004499')],
              foreground=[('active', 'white'), ('pressed', 'white')],
              bordercolor=[('active', '#0055B0'), ('pressed', '#004499')])

    # 配置进度条样式
    style.configure('TProgressbar',
                    background=mac_accent,
                    troughcolor=mac_secondary,
                    bordercolor=mac_secondary,
                    thickness=6,
                    relief='flat')

    # 设置窗口尺寸和位置
    window_width = 750
    window_height = 550
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    root.minsize(650, 500)

    # 显示窗口
    root.deiconify()

    # 初始化全局变量
    status_var = tk.StringVar()
    bom_var = tk.StringVar()
    sub_var = tk.StringVar()

    # 加载配置
    config = load_config()

    # 打印配置中的替代料表路径，用于调试
    print("配置中的替代料表路径:", config.get('default_sub_path', ''))
    logging.info(f"配置中的替代料表路径: {config.get('default_sub_path', '')}")

    # 如果有默认替代料表路径，自动加载
    if config.get('default_sub_path') and os.path.exists(config.get('default_sub_path')):
        print(f"自动加载替代料表路径: {config['default_sub_path']}")
        logging.info(f"自动加载替代料表路径: {config['default_sub_path']}")
        sub_var.set(config['default_sub_path'])
    else:
        if not config.get('default_sub_path'):
            print("配置中没有替代料表路径")
            logging.info("配置中没有替代料表路径")
        elif not os.path.exists(config.get('default_sub_path')):
            print(f"替代料表路径不存在: {config.get('default_sub_path')}")
            logging.info(f"替代料表路径不存在: {config.get('default_sub_path')}")

    # 主容器
    main_frame = ttk.Frame(root, padding=(20, 15))  # 减小上下padding，原来是(30, 25)
    main_frame.pack(fill='both', expand=True)

    # 顶部区域
    top_frame = ttk.Frame(main_frame)
    top_frame.pack(fill='x', pady=(0, 10))  # 减小顶部和内容之间的间距，原来是15

    # 标题和工具栏
    title_frame = ttk.Frame(top_frame)
    title_frame.pack(fill='x')

    # 标题
    title_label = ttk.Label(title_frame,
                          text="BOM替代料工具",
                           font=(default_font_family, 16, 'bold'),  # 减小字体大小，原来是18
                           foreground="#333333")
    title_label.pack(side='left')

    # 工具栏/按钮区域
    toolbar_frame = ttk.Frame(title_frame)
    toolbar_frame.pack(side='right')

    # 设置按钮
    settings_button = ttk.Button(toolbar_frame,
                               text="设置",
                               command=show_header_config,
                               width=6)  # 减小按钮宽度，原来是8
    settings_button.pack(side='right', padx=3)  # 减小按钮间距，原来是5

    # 检查更新按钮
    update_button = ttk.Button(toolbar_frame,
                             text="检查更新",
                             command=lambda: update_manager.check_updates_manually(),
                             width=8)  # 由于文字较长，使用稍大的宽度
    update_button.pack(side='right', padx=3)

    # 帮助按钮
    help_button = ttk.Button(toolbar_frame,
                            text="帮助",
                           command=show_help,
                            width=6)  # 减小按钮宽度，原来是8
    help_button.pack(side='right', padx=3)  # 减小按钮间距，原来是5

    # 内容区域
    content_frame = ttk.Frame(main_frame)
    content_frame.pack(fill='both', expand=True)

    # 文件选择区域
    file_section = ttk.Frame(content_frame)
    file_section.pack(fill='x', pady=(0, 10))  # 减小文件选择区域底部间距，原来是15

    # 区域标题
    section_label = ttk.Label(file_section,
                             text="文件选择",
                             font=(default_font_family, 11, 'bold'),  # 减小字体大小，原来是13
                             foreground="#333333")
    section_label.pack(anchor='w', pady=(0, 3))  # 减小标题下方间距，原来是5

    # 文件选择容器
    file_container = ttk.Frame(file_section, padding=8, relief="solid", borderwidth=1)  # 减小内部padding，原来是10
    file_container.pack(fill='x')

    # BOM文件选择
    bom_file_frame = ttk.Frame(file_container)
    bom_file_frame.pack(fill='x', pady=(0, 6))  # 减小行间距，原来是8

    ttk.Label(bom_file_frame,
             text="BOM文件:",
             font=(default_font_family, 10),  # 移除bold
             width=8).pack(side='left', padx=(0, 8))  # 减小标签宽度和间距

    # 文件路径显示
    bom_entry = ttk.Entry(bom_file_frame, textvariable=bom_var, width=50)
    bom_entry.pack(side='left', fill='x', expand=True, padx=(0, 8))  # 减小间距

    # 选择文件按钮
    bom_button = ttk.Button(bom_file_frame,
                          text="浏览...",
                          command=lambda: select_file(bom_var, 'xlsx'),
                          width=6)  # 减小按钮宽度
    bom_button.pack(side='right')

    # 替代料表选择
    sub_file_frame = ttk.Frame(file_container)
    sub_file_frame.pack(fill='x')

    ttk.Label(sub_file_frame,
             text="替代料表:",
             font=(default_font_family, 10),  # 移除bold
             width=8).pack(side='left', padx=(0, 8))  # 减小标签宽度和间距

    # 文件路径显示
    sub_entry = ttk.Entry(sub_file_frame, textvariable=sub_var, width=50)
    sub_entry.pack(side='left', fill='x', expand=True, padx=(0, 8))  # 减小间距

    # 选择文件按钮
    sub_button = ttk.Button(sub_file_frame,
                          text="浏览...",
                          command=lambda: select_file(sub_var, 'xlsx', is_sub_file=True),
                          width=6)  # 减小按钮宽度
    sub_button.pack(side='right')

    # 操作按钮框架
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill='x', pady=(0, 8))  # 减小操作按钮下方间距，原来是10

    # 开始处理按钮
    start_button = ttk.Button(button_frame, text='开始处理',
                            command=lambda: Thread(target=process_files).start(),
                            style='Primary.TButton',
                            width=12)  # 减小按钮宽度，原来是15
    start_button.pack(side='left')

    # 状态提示区域
    status_frame = ttk.LabelFrame(main_frame, text=' 处理进度 ', padding=(12, 8))  # 减小padding，原来是(15, 10)
    status_frame.pack(fill='both', expand=True)

    # 进度条和百分比容器
    progress_container = ttk.Frame(status_frame)
    progress_container.pack(fill='x', pady=(0, 6))  # 减小进度条下方间距，原来是8

    # 进度条 - 改进样式
    global progress
    progress = ttk.Progressbar(progress_container, orient='horizontal', mode='determinate', length=400)
    progress.pack(side='left', fill='x', expand=True, padx=(0, 8))

    # 百分比标签
    global progress_percent
    progress_percent = ttk.Label(progress_container, text='0%', width=5)
    progress_percent.pack(side='left')

    # 添加滚动条和文本区域
    status_container = ttk.Frame(status_frame)
    status_container.pack(fill='both', expand=True)

    # 滚动条
    status_scroll = ttk.Scrollbar(status_container)
    status_scroll.pack(side='right', fill='y')

    # 状态文本框 - 改进字体和背景
    global status_text
    status_text = tk.Text(status_container, wrap=tk.WORD, height=15,  # 增加文本框高度，原来是10
                          yscrollcommand=status_scroll.set,
                         font=('微软雅黑', 9),
                         background=mac_secondary,  # 使用mac_secondary变量，而不是secondary_color
                         padx=10, pady=10)
    status_text.pack(fill='both', expand=True)
    status_scroll.config(command=status_text.yview)

    # 初始状态
    update_status("就绪，请选择文件并点击\"开始处理\"按钮")

    # 底部信息栏
    footer_frame = ttk.Frame(main_frame)
    footer_frame.pack(fill='x', pady=(10, 0))

    # 作者信息
    author_label = ttk.Label(footer_frame,
                           text="开发者: 小航 | 联系: XiaoHang_Sky",
                           font=('微软雅黑', 8))
    author_label.pack(side='left')

    # 日期信息
    date_label = ttk.Label(footer_frame,
                         text="2025.5.12",
                         font=('微软雅黑', 8))
    date_label.pack(side='right')

    # 配置文本标签样式
    status_text.tag_configure('title', font=('微软雅黑', 11, 'bold'), foreground='#0078D4')
    status_text.tag_configure('separator', foreground='#808080')
    status_text.tag_configure('success', foreground='#107C10', font=('微软雅黑', 10, 'bold'))
    status_text.tag_configure('item', foreground='#000000')
    status_text.tag_configure('subtitle', foreground='#0078D4', font=('微软雅黑', 9, 'bold'))

    # 改进进度条样式
    style.configure("TProgressbar",
                   thickness=16,
                   borderwidth=0,
                   background='#0078D4',
                   troughcolor='#E6E6E6')

    root.mainloop()

def select_file(var, ext, is_sub_file=False):
    """选择文件

    Args:
        var: 存储文件路径的变量
        ext: 文件扩展名
        is_sub_file: 是否是替代料文件
    """
    # 加载配置
    config = load_config()

    # 设置初始目录
    if is_sub_file:
        # 替代料文件使用保存的默认路径或用户主目录
        initial_dir = os.path.dirname(config['default_sub_path']) if config['default_sub_path'] else os.path.expanduser("~")
    else:
        # BOM文件使用上次打开的目录，如无则使用桌面或用户主目录
        if config['last_bom_dir'] and os.path.exists(config['last_bom_dir']):
            initial_dir = config['last_bom_dir']
        else:
            # 尝试使用桌面路径
            desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            # 确保桌面路径存在，不存在则使用用户主目录
            initial_dir = desktop_dir if os.path.exists(desktop_dir) else os.path.expanduser("~")

    filename = filedialog.askopenfilename(
        filetypes=[('Excel文件', f'*.{ext}')],
        initialdir=initial_dir  # 使用设置的初始目录
    )

    if filename:
        var.set(filename)

        # 如果是BOM文件，保存目录到配置
        if not is_sub_file:
            config['last_bom_dir'] = os.path.dirname(filename)
            save_config(config)

        # 如果是替代料文件，直接设置为默认路径并保存
        if is_sub_file:
            # 如果路径变更，更新配置
            if filename != config['default_sub_path']:
                config['default_sub_path'] = filename
                save_config(config)
                print(f"已自动将 {filename} 设置为默认替代料表路径")
                logging.info(f"已自动将 {filename} 设置为默认替代料表路径")

def update_progress(value):
    global root
    if root and root.winfo_exists():
        root.after(0, lambda: [
            progress.config(value=value),
            progress_percent.config(text=f'{int(value)}%'),
            root.update_idletasks()
        ])

def update_status(message, color=None):
    """更新状态文本

    Args:
        message: 状态消息
        color: 文本颜色（可选）
    """
    global status_text, update_manager
    if status_text:
        status_text.config(state=tk.NORMAL)  # 临时允许编辑
        status_text.delete(1.0, tk.END)  # 清除现有内容

        # 使用标签和颜色美化文本显示
        lines = message.split('\n')
        for line in lines:
            if any(marker in line for marker in ["✅", "📊", "📋", "🔄"]):
                # 部分标题使用蓝色粗体
                status_text.insert(tk.END, line + "\n", 'title')
            elif line.startswith("-"):
                # 分隔线使用灰色
                status_text.insert(tk.END, line + "\n", 'separator')
            elif "处理完成" in line:
                # 完成提示使用绿色
                status_text.insert(tk.END, line + "\n", 'success')
            elif line.strip().startswith("•"):
                # 统计项目使用黑色
                status_text.insert(tk.END, line + "\n", 'item')
            elif "物料" in line and ":" in line:
                # 物料标题使用蓝色
                status_text.insert(tk.END, line + "\n", 'subtitle')
            elif color:
                # 使用指定颜色
                # 创建临时标签
                tag_name = f"color_{color.replace('#', '')}"
                if not tag_name in status_text.tag_names():
                    status_text.tag_configure(tag_name, foreground=color)
                status_text.insert(tk.END, line + "\n", tag_name)
            else:
                # 普通文本
                status_text.insert(tk.END, line + "\n")

        status_text.see(1.0)  # 滚动到顶部
        status_text.config(state=tk.DISABLED)  # 恢复只读状态

        # 更新窗口
        if root and root.winfo_exists():
            root.update_idletasks()

    # 如果更新管理器存在，更新其状态栏方法
    if 'update_manager' in globals() and update_manager:
        # 覆盖UpdateManager类的_update_status方法
        update_manager._update_status = lambda msg, clr=None: update_status(msg, clr)

def show_help():
    """显示使用帮助对话框，带有标签页和格式化文本"""
    help_window = tk.Toplevel()
    help_window.withdraw()  # 先隐藏窗口，避免闪烁
    help_window.title("BOM替代料工具 - 使用帮助")

    # 设置窗口尺寸和位置
    window_width = 700
    window_height = 550

    # 计算窗口居中位置
    screen_width = help_window.winfo_screenwidth()
    screen_height = help_window.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)

    # 设置窗口位置
    help_window.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

    # 设置窗口图标（如果可用）
    try:
        icon_path = resource_path('DSC01.ico')
        if os.path.exists(icon_path):
            help_window.iconbitmap(icon_path)
    except Exception:
        pass

    # 设置样式
    style = ttk.Style()
    style.configure("Tab.TFrame", background="#f5f5f5")

    # 创建主框架
    main_frame = ttk.Frame(help_window, padding=10)
    main_frame.pack(fill='both', expand=True)

    # 创建标签页控件
    tab_control = ttk.Notebook(main_frame)

    # 创建各个标签页
    tab_overview = ttk.Frame(tab_control, style="Tab.TFrame")
    tab_updates = ttk.Frame(tab_control, style="Tab.TFrame")
    tab_usage = ttk.Frame(tab_control, style="Tab.TFrame")
    tab_contact = ttk.Frame(tab_control, style="Tab.TFrame")

    tab_control.add(tab_overview, text=" 概览 ")
    tab_control.add(tab_updates, text=" 更新说明 ")
    tab_control.add(tab_usage, text=" 使用方法 ")
    tab_control.add(tab_contact, text=" 技术支持 ")

    tab_control.pack(fill="both", expand=True)

    # === 概览标签页内容 ===
    overview_frame = ttk.Frame(tab_overview, padding=15)
    overview_frame.pack(fill='both', expand=True)

    # 标题
    title_label = ttk.Label(overview_frame,
                         text="BOM替代料工具",
                         font=('微软雅黑', 16, 'bold'))
    title_label.pack(anchor='w', pady=(0, 10))

    # 当前版本
    version_frame = ttk.Frame(overview_frame)
    version_frame.pack(fill='x', pady=(0, 15))

    version_label = ttk.Label(version_frame,
                            text="当前版本：v2.5",
                            font=('微软雅黑', 10, 'bold'))
    version_label.pack(side='left')

    date_label = ttk.Label(version_frame,
                         text="发布日期：2025年5月12日",
                         font=('微软雅黑', 10))
    date_label.pack(side='right')

    ttk.Separator(overview_frame, orient='horizontal').pack(fill='x', pady=10)

    # 工具概述
    overview_text = tk.Text(overview_frame, wrap=tk.WORD, height=20,
                          font=('微软雅黑', 10), background="#f9f9f9",
                          padx=15, pady=15, borderwidth=1, relief="solid")
    overview_text.pack(fill='both', expand=True)

    overview_content = """BOM替代料工具是一款专业的电子制造业辅助软件，专为解决电子BOM（物料清单）管理过程中的替代料处理而设计。

本工具能自动识别并处理BOM表中的替代料关系，显著提高物料管理效率和准确性。主要功能包括：

• BOM自动重构 - 生成带完整替代料标识的新BOM
• 相同料号智能合并 - 自动合并相同物料的多个条目
• 位号统计精确化 - 准确计算每个物料的位号数量
• 表头配置功能 - 支持自定义BOM表头字段名称

推荐使用场景：
- PCB生产物料清单处理
- BOM标准化与一致性检查
- 物料替代关系管理
- 供应链物料优化
"""

    overview_text.insert(tk.END, overview_content)
    overview_text.config(state=tk.DISABLED)

    # === 更新说明标签页内容 ===
    updates_frame = ttk.Frame(tab_updates, padding=15)
    updates_frame.pack(fill='both', expand=True)

    # 创建滚动条
    updates_scroll = ttk.Scrollbar(updates_frame)
    updates_scroll.pack(side='right', fill='y')

    updates_text = tk.Text(updates_frame, wrap=tk.WORD, yscrollcommand=updates_scroll.set,
                         font=('微软雅黑', 10), background="#f9f9f9",
                         padx=15, pady=15, borderwidth=1, relief="solid")
    updates_text.pack(fill='both', expand=True)
    updates_scroll.config(command=updates_text.yview)

    # 配置样式标签
    updates_text.tag_configure('version', font=('微软雅黑', 12, 'bold'), foreground='#0078D4')
    updates_text.tag_configure('date', font=('微软雅黑', 10), foreground='#666666')
    updates_text.tag_configure('category', font=('微软雅黑', 10, 'bold'), foreground='#107C10')
    updates_text.tag_configure('bullet', foreground='#0078D4')
    updates_text.tag_configure('separator', foreground='#CCCCCC')

    # V2.5 更新说明
    updates_text.insert(tk.END, "V2.5 更新说明\n", 'version')
    updates_text.insert(tk.END, "发布日期：2025年5月12日\n\n", 'date')

    updates_text.insert(tk.END, "问题修复\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "修复表头大小写不敏感问题，支持不同大小写的表头名称（如Description、DESCRIPTION等）\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "修复处理浮点数类型数据时的错误，增强数据类型兼容性\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "优化表头匹配逻辑，自动适应实际表头的大小写\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "增强程序健壮性，即使某些列不存在也能继续运行\n\n")

    # 分隔线
    updates_text.insert(tk.END, "-" * 60 + "\n\n", 'separator')

    # V2.4 更新说明
    updates_text.insert(tk.END, "V2.4 更新说明\n", 'version')
    updates_text.insert(tk.END, "发布日期：2025年4月16日\n\n", 'date')

    updates_text.insert(tk.END, "主要更新\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "修复已知bug\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "增加更新检测功能\n")


    # 分隔线
    updates_text.insert(tk.END, "-" * 60 + "\n\n", 'separator')

    # V2.3 更新说明
    updates_text.insert(tk.END, "V2.3 更新说明\n", 'version')
    updates_text.insert(tk.END, "发布日期：2025年3月19日\n\n", 'date')

    updates_text.insert(tk.END, "主要更新\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "全新界面设计，操作更简单直观\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "支持自定义替代料的高亮颜色\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "优化大文件处理速度，运行更快\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "改进错误提示，更容易理解问题\n\n")

    # 分隔线
    updates_text.insert(tk.END, "-" * 60 + "\n\n", 'separator')

    # V2.2 更新说明
    updates_text.insert(tk.END, "V2.2 更新说明\n", 'version')
    updates_text.insert(tk.END, "发布日期：2025年3月15日\n\n", 'date')

    updates_text.insert(tk.END, "主要更新\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "新增表头配置功能，支持自定义BOM表头\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "自动保存常用设置，减少重复操作\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "优化替代料识别准确度\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "实时显示处理进度和状态\n\n")

    # 分隔线
    updates_text.insert(tk.END, "-" * 60 + "\n\n", 'separator')

    # V2.1 更新说明
    updates_text.insert(tk.END, "V2.1 更新说明\n", 'version')
    updates_text.insert(tk.END, "发布日期：2025年3月14日\n\n", 'date')

    updates_text.insert(tk.END, "问题修复\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "修复了项目信息行格式保存和恢复的问题\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "优化了Excel文件读写性能\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "改进了替代料编号逻辑，避免误判\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "修复了原始BOM中item号相同但位号不同时的误判问题\n\n")

    updates_text.insert(tk.END, "功能优化\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "新增自动识别和保留项目信息行功能\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "优化了Excel格式处理，支持更多样式属性\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "新增BOM预处理功能，对原始BOM的item进行顺序编号\n\n")

    # 分隔线
    updates_text.insert(tk.END, "-" * 60 + "\n\n", 'separator')

    # V2.0 更新说明
    updates_text.insert(tk.END, "V2.0 更新说明\n", 'version')
    updates_text.insert(tk.END, "发布日期：2025年3月8日\n\n", 'date')

    updates_text.insert(tk.END, "核心功能增强\n", 'category')
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "相同料号智能合并\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "Item序号连续化\n")
    updates_text.insert(tk.END, "• ", 'bullet')
    updates_text.insert(tk.END, "位号统计精确化\n\n")

    updates_text.config(state=tk.DISABLED)

    # === 使用方法标签页内容 ===
    usage_frame = ttk.Frame(tab_usage, padding=15)
    usage_frame.pack(fill='both', expand=True)

    # 创建滚动条
    usage_scroll = ttk.Scrollbar(usage_frame)
    usage_scroll.pack(side='right', fill='y')

    usage_text = tk.Text(usage_frame, wrap=tk.WORD, yscrollcommand=usage_scroll.set,
                       font=('微软雅黑', 10), background="#f9f9f9",
                       padx=15, pady=15, borderwidth=1, relief="solid")
    usage_text.pack(fill='both', expand=True)
    usage_scroll.config(command=usage_text.yview)

    # 配置样式标签
    usage_text.tag_configure('section', font=('微软雅黑', 12, 'bold'), foreground='#0078D4')
    usage_text.tag_configure('subsection', font=('微软雅黑', 11, 'bold'), foreground='#333333')
    usage_text.tag_configure('important', font=('微软雅黑', 10, 'bold'), background='#FFF4CE')
    usage_text.tag_configure('bullet', foreground='#0078D4')
    usage_text.tag_configure('step', font=('微软雅黑', 10, 'bold'), foreground='#107C10')
    usage_text.tag_configure('note', font=('微软雅黑', 9), foreground='#666666', background='#E6F2FF')

    # 基本操作部分
    usage_text.insert(tk.END, "基本操作\n", 'section')
    usage_text.insert(tk.END, "\n1. ", 'step')
    usage_text.insert(tk.END, '选择BOM文件：点击"选择文件"按钮，选择需要处理的BOM文件\n')
    usage_text.insert(tk.END, "2. ", 'step')
    usage_text.insert(tk.END, '选择替代料表：点击"选择文件"按钮，选择包含替代料信息的Excel表格\n')
    usage_text.insert(tk.END, "3. ", 'step')
    usage_text.insert(tk.END, '点击"开始处理"：系统会自动分析BOM文件，添加替代料并生成新的BOM文件\n\n')

    # 表头设置部分
    usage_text.insert(tk.END, "表头设置\n", 'section')
    usage_text.insert(tk.END, "\n如果您的BOM文件或替代料表使用的是自定义表头，请使用表头设置功能：\n\n")
    usage_text.insert(tk.END, "1. ", 'step')
    usage_text.insert(tk.END, '点击主界面上的"表头设置"按钮\n')
    usage_text.insert(tk.END, "2. ", 'step')
    usage_text.insert(tk.END, "在弹出的对话框中，为各字段配置对应的表头名称：\n")

    field_descriptions = [
        "Item编号字段：对应物料的序号字段",
        "物料编号字段：对应物料的唯一标识编号",
        "零件字段：对应物料的零件名称",
        "位号字段：对应物料在PCB上的位置标识",
        "数量字段：对应物料的数量",
        "描述字段：对应物料的描述信息",
        "制造商料号字段：对应物料的制造商料号",
        "制造商字段：对应物料的制造商名称",
        "替代料属性字段：对应替代料表中的属性字段，用于识别替代料组"
    ]

    for desc in field_descriptions:
        usage_text.insert(tk.END, "   • ", 'bullet')
        usage_text.insert(tk.END, desc + "\n")

    usage_text.insert(tk.END, "\n3. ", 'step')
    usage_text.insert(tk.END, '点击"保存配置"按钮保存设置\n\n')

    # 替代料表格式要求
    usage_text.insert(tk.END, "替代料表格式要求\n", 'section')
    usage_text.insert(tk.END, "\n", 'important')
    usage_text.insert(tk.END, "替代料表必须满足以下格式要求：\n", 'important')
    usage_text.insert(tk.END, "\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "替代料表必须包含与BOM相同的物料编号字段\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "【仅替代料表】必须包含属性字段(attribute)，用于识别替代料组关系\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "同一替代组内的物料具有相同的attribute值\n\n")

    # 输出结果
    usage_text.insert(tk.END, "输出结果\n", 'section')
    usage_text.insert(tk.END, "\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, '程序会在原BOM文件所在目录生成以"_替代料"为后缀的新Excel文件\n')
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "替代料会以黄色底色高亮显示\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "处理完成后，会显示详细的处理统计信息\n\n")

    # 注意事项
    usage_text.insert(tk.END, "注意事项\n", 'section')
    usage_text.insert(tk.END, "\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "处理大文件时可能需要较长时间，请耐心等待\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "表头关键字匹配不区分大小写，但建议保持与Excel表头完全一致\n")
    usage_text.insert(tk.END, "• ", 'bullet')
    usage_text.insert(tk.END, "对于经常使用的替代料表，可设置为默认路径以简化操作\n")

    usage_text.config(state=tk.DISABLED)

    # === 技术支持标签页内容 ===
    contact_frame = ttk.Frame(tab_contact, padding=15)
    contact_frame.pack(fill='both', expand=True)

    support_label = ttk.Label(contact_frame,
                            text="技术支持",
                            font=('微软雅黑', 14, 'bold'))
    support_label.pack(anchor='w', pady=(0, 15))

    contact_text = tk.Text(contact_frame, wrap=tk.WORD, height=10,
                         font=('微软雅黑', 10), background="#f9f9f9",
                         padx=15, pady=15, borderwidth=1, relief="solid")
    contact_text.pack(fill='both', expand=True)

    # 添加联系信息
    contact_text.insert(tk.END, "如果您在使用过程中遇到任何问题，或者有功能改进建议，请通过以下方式联系我们：\n\n")
    contact_text.insert(tk.END, "开发者：小航\n")
    contact_text.insert(tk.END, "联系方式：XiaoHang_Sky（微信）\n\n")
    contact_text.insert(tk.END, "我们将尽快回复您的问题并提供技术支持。\n\n")
    contact_text.insert(tk.END, "感谢您使用BOM替代料工具！")

    contact_text.config(state=tk.DISABLED)

    # 底部的关闭按钮
    close_btn = ttk.Button(main_frame, text="关闭", command=help_window.destroy, width=12)
    close_btn.pack(pady=10)

    # 窗口置顶
    help_window.transient(root)
    help_window.grab_set()

    # 准备好所有内容后再显示窗口，避免闪烁
    help_window.deiconify()

    root.wait_window(help_window)

def count_references(reference_text):
    """
    计算位号字符串中的有效位号数量

    Args:
        reference_text: 位号字符串，可能包含用逗号分隔的多个位号

    Returns:
        int: 有效位号的数量
    """
    if pd.isna(reference_text) or not str(reference_text).strip():
        return 0

    # 拆分位号并去除空位号
    references = [ref.strip() for ref in str(reference_text).split(',') if ref.strip()]
    return len(references)

def show_custom_error(title, message, parent=None):
    """显示自定义错误对话框

    Args:
        title: 对话框标题
        message: 错误信息
        parent: 父窗口，默认为root
    """
    if parent is None:
        parent = root

    # 创建对话框
    error_dialog = tk.Toplevel(parent)
    error_dialog.title(title)
    error_dialog.transient(parent)  # 设置为父窗口的子窗口
    error_dialog.grab_set()  # 模态对话框

    # 设置窗口尺寸和位置
    window_width = 450
    window_height = 250
    screen_width = error_dialog.winfo_screenwidth()
    screen_height = error_dialog.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    error_dialog.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
    error_dialog.minsize(450, 200)

    # 主框架
    main_frame = ttk.Frame(error_dialog, padding=15)
    main_frame.pack(fill='both', expand=True)

    # 错误图标和标题框架
    header_frame = ttk.Frame(main_frame)
    header_frame.pack(fill='x', pady=(0, 15))

    # 错误图标（使用文本代替）
    error_icon = ttk.Label(header_frame, text="⚠", font=('Arial', 24, 'bold'), foreground='#FF0000')
    error_icon.pack(side='left', padx=(0, 10))

    # 错误标题
    title_label = ttk.Label(header_frame, text=title, font=('\u5fae\u8f6f\u96c5\u9ed1', 12, 'bold'))
    title_label.pack(side='left')

    # 错误信息框架
    message_frame = ttk.Frame(main_frame)
    message_frame.pack(fill='both', expand=True, pady=(0, 15))

    # 错误信息文本框
    message_text = tk.Text(message_frame, wrap=tk.WORD, height=6, width=50,
                          font=('\u5fae\u8f6f\u96c5\u9ed1', 10),
                          background='#F5F5F7', relief='flat', padx=10, pady=10)
    message_text.pack(fill='both', expand=True)
    message_text.insert(tk.END, message)
    message_text.config(state='disabled')  # 设置为只读

    # 添加滚动条
    scrollbar = ttk.Scrollbar(message_text, command=message_text.yview)
    scrollbar.pack(side='right', fill='y')
    message_text.config(yscrollcommand=scrollbar.set)

    # 按钮框架
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill='x', pady=(0, 5))

    # 确定按钮
    ok_button = ttk.Button(button_frame, text="确定", width=10, command=error_dialog.destroy)
    ok_button.pack(side='right')

    # 设置默认焦点并绑定Enter键
    ok_button.focus_set()
    error_dialog.bind("<Return>", lambda event: error_dialog.destroy())
    error_dialog.bind("<Escape>", lambda event: error_dialog.destroy())

    # 等待用户关闭对话框
    parent.wait_window(error_dialog)

def translate_error_to_chinese(error):
    """将英文错误信息转换为中文错误信息"""
    # 常见错误信息的中英文映射
    error_mapping = {
        # 文件操作相关错误
        "Permission denied": "文件权限不足，请以管理员身份运行或检查文件权限",
        "File not found": "找不到指定的文件，请检查文件路径是否正确",
        "Invalid file": "无效的文件，请检查文件格式是否正确",
        "Failed to open": "打开文件失败，请确保文件未被其他程序占用",
        "Cannot read": "无法读取文件，请检查文件是否损坏或格式不正确",
        "Cannot write": "无法写入文件，请检查文件权限或磁盘空间",

        # Excel文件相关错误
        "Excel file format": "Excel文件格式错误，请使用标准的Excel格式(.xlsx或.xls)",
        "XLRDError": "不支持的Excel文件格式，请使用.xlsx或.xls格式",
        "EmptyDataError": "Excel文件内容为空，请检查文件是否有数据",
        "No sheet named": "找不到指定的工作表，请检查Excel文件",
        "Sheet index out of range": "工作表索引超出范围，请检查Excel文件",

        # 系统资源相关错误
        "Memory error": "内存不足，请关闭其他程序或增加系统内存",
        "Disk full": "磁盘空间不足，请清理磁盘空间",

        # 网络相关错误
        "Timeout": "网络连接超时，请检查网络连接或重试",
        "Connection error": "网络连接错误，请检查网络连接",
        "ConnectionError": "网络连接错误，请检查网络连接",
        "ConnectionRefusedError": "连接被拒绝，请检查网络设置或防火墙",

        # 编码相关错误
        "UnicodeDecodeError": "文件编码错误，请使用UTF-8或GBK编码保存文件",
        "UnicodeEncodeError": "文本编码错误，可能包含不支持的字符",

        # Python内部错误
        "KeyError": "程序内部错误：找不到指定的键值，请检查数据格式",
        "IndexError": "程序内部错误：索引超出范围，请检查数据格式",
        "TypeError": "程序内部错误：类型错误，请检查数据格式",
        "ValueError": "程序内部错误：值错误，请检查数据格式",
        "AttributeError": "程序内部错误：属性错误，请检查数据格式",

        # 模块相关错误
        "ImportError": "程序内部错误：导入模块失败，请重新安装程序",
        "ModuleNotFoundError": "缺少必要的模块，请重新安装程序",
        "No module named": "缺少必要的模块，请重新安装程序",

        # 系统错误
        "OSError": "操作系统错误，请检查文件权限或磁盘空间",
        "FileNotFoundError": "找不到指定的文件，请检查文件路径是否正确",
        "PermissionError": "文件权限不足，请以管理员身份运行或检查文件权限",
        "FileExistsError": "文件已存在，请尝试使用其他文件名",

        # 其他错误
        "NotImplementedError": "功能尚未实现，请等待后续版本",
        "RuntimeError": "运行时错误，请重新启动程序或联系开发者",
        "Exception": "程序异常，请重新启动程序或联系开发者"
    }

    # 检查错误信息是否包含已知的错误模式
    error_str = str(error)
    for eng_error, cn_error in error_mapping.items():
        if eng_error in error_str:
            return cn_error

    # 如果没有匹配到已知错误，尝试根据错误类型提供通用提示
    if "pandas" in error_str:
        return f"数据处理错误：{error_str}，请检查Excel文件格式是否正确"
    elif "openpyxl" in error_str:
        return f"Excel文件处理错误：{error_str}，请检查Excel文件是否损坏"
    elif "requests" in error_str:
        return f"网络请求错误：{error_str}，请检查网络连接"

    # 如果没有匹配到任何已知错误，返回原始错误信息
    return f"程序错误：{error_str}"

def process_files():
    try:
        # 记录开始时间
        start_time = time.time()

        # 获取全局变量
        global bom_var, sub_var, status_var, progress, progress_percent

        # 确保变量已初始化
        if not hasattr(bom_var, 'get') or not hasattr(sub_var, 'get'):
            tkinter.messagebox.showerror('错误', '请先选择文件')
            return

        # 初始化进度条
        progress['maximum'] = 100
        update_progress(0)
        update_status("开始处理文件...")

        # 更新状态
        update_status('正在处理中...')

        # 获取选择的文件路径
        bom_path = bom_var.get()
        sub_path = sub_var.get()

        if not bom_path or not sub_path:
            tkinter.messagebox.showerror('错误', '请先选择BOM文件和替代料表')
            return

        # 加载配置中的表头映射
        config = load_config()
        bom_header_mapping = config['bom_header_mapping']  # BOM表头映射
        sub_header_mapping = config['sub_header_mapping']  # 替代料表表头映射

        # 获取BOM文件必需列表头
        required_bom_columns = [
            bom_header_mapping['item'],           # Item编号字段
            bom_header_mapping['pn'],             # 物料编号字段
            bom_header_mapping['part'],           # 零件字段
            bom_header_mapping['reference'],      # 位号字段
            bom_header_mapping['quantity'],       # 数量字段
            bom_header_mapping['description'],    # 描述字段
            bom_header_mapping['mfr_pn'],         # 制造商料号字段
            bom_header_mapping['manufacturer']    # 制造商字段
        ]

        # 获取替代料表必需列表头
        required_sub_columns = [
            sub_header_mapping['pn'],             # 物料编号字段
            sub_header_mapping['attribute']       # 替代料属性字段
        ]

        # 识别项目信息行
        logging.info("开始识别项目信息行")
        update_status('正在识别项目信息行...')

        # 读取原始Excel文件以获取格式信息
        original_wb = openpyxl.load_workbook(bom_path)
        original_ws = original_wb.active

        # 找到第一个包含必需列的行
        header_row = None

        for row_idx, row in enumerate(original_ws.iter_rows(min_row=1, max_row=original_ws.max_row), 1):
            row_values = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
            # 检查是否至少有一半的必需列存在于当前行
            matches = sum(1 for col in required_bom_columns if any(col.lower() == val.lower() for val in row_values))
            if matches >= len(required_bom_columns) / 2:  # 如果有至少一半的列匹配
                header_row = row_idx
                # 记录实际找到的表头，用于后续处理
                found_headers = {val.lower(): val for val in row_values if val}
                # 更新last_used_header_mapping，记录实际使用的表头
                for key, expected_header in bom_header_mapping.items():
                    if expected_header.lower() in found_headers:
                        config['last_used_header_mapping'][key] = found_headers[expected_header.lower()]
                save_config(config)
                break

        if header_row is None:
            raise ValueError("无法在BOM文件中找到必需列，请检查表头配置是否正确")

        # 保存项目信息行
        project_info_rows = []
        for row_idx in range(1, header_row):
            row_data = {}
            for col_idx, cell in enumerate(original_ws[row_idx], 1):
                # 保存样式属性而不是样式对象
                row_data[col_idx] = {
                    'value': cell.value,
                    'font_name': cell.font.name,
                    'font_size': cell.font.size,
                    'font_bold': cell.font.bold,
                    'fill_type': cell.fill.fill_type,
                    'fill_color': cell.fill.start_color.rgb if cell.fill.start_color else None,
                    'border_left': cell.border.left.style if cell.border.left else None,
                    'border_right': cell.border.right.style if cell.border.right else None,
                    'border_top': cell.border.top.style if cell.border.top else None,
                    'border_bottom': cell.border.bottom.style if cell.border.bottom else None,
                    'alignment_horizontal': cell.alignment.horizontal,
                    'alignment_vertical': cell.alignment.vertical,
                    'number_format': cell.number_format
                }
            project_info_rows.append(row_data)

        # 更新进度
        update_progress(10)

        # 使用pandas读取BOM文件，跳过项目信息行
        logging.info(f"读取BOM文件: {bom_path}，跳过前 {header_row-1} 行")
        bom_df = pd.read_excel(bom_path, dtype={bom_header_mapping['item']: str}, skiprows=header_row-1)
        logging.info(f"BOM文件列: {list(bom_df.columns)}")

        # 单独读取替代料表，不应用项目信息行的跳过
        logging.info(f"读取替代料表: {sub_path}")
        try:
            sub_df = pd.read_excel(sub_path, dtype={sub_header_mapping['pn']: str})
            logging.info(f"替代料表列: {list(sub_df.columns)}")
        except Exception as e:
            error_msg = translate_error_to_chinese(e)
            logging.error(f"读取替代料表失败: {e}")

            # 使用自定义错误对话框
            error_details = f"读取替代料表时出错：\n\n{error_msg}\n\n请检查文件格式是否正确。"
            show_custom_error('读取文件错误', error_details)
            return

        # 确保BOM文件表头字段存在（不区分大小写）
        missing_bom_fields = []
        # 创建列名的小写映射，用于不区分大小写的匹配
        columns_lower = {col.lower(): col for col in bom_df.columns}

        for field, header in bom_header_mapping.items():
            # 检查表头是否存在（不区分大小写）
            if header.lower() in columns_lower:
                # 如果存在但大小写不同，使用实际的列名替换配置中的列名
                actual_column = columns_lower[header.lower()]
                if actual_column != header:
                    logging.info(f"表头大小写不同，使用实际列名: '{actual_column}' 替代 '{header}'")
                    bom_header_mapping[field] = actual_column
            else:
                missing_bom_fields.append(header)
                # 对于Description列，提供特殊处理
                if field == 'description':
                    tkinter.messagebox.showwarning('警告', f'BOM文件中未找到表头 "{header}"，请检查表头配置')
                else:
                    tkinter.messagebox.showwarning('警告', f'BOM文件中未找到表头 "{header}"，请检查表头配置')

        if missing_bom_fields:
            logging.warning(f"BOM文件缺少以下字段: {missing_bom_fields}")

            # 如果缺少Description列，添加一个空列以避免后续处理错误
            if bom_header_mapping['description'] not in bom_df.columns:
                bom_df[bom_header_mapping['description']] = ""
                logging.info(f"已添加空的Description列: {bom_header_mapping['description']}")

        # 确保替代料表表头字段存在（不区分大小写）
        missing_sub_fields = {'required': [], 'optional': []}
        # 创建列名的小写映射，用于不区分大小写的匹配
        sub_columns_lower = {col.lower(): col for col in sub_df.columns}

        for field, header in sub_header_mapping.items():
            # 检查表头是否存在（不区分大小写）
            if header.lower() in sub_columns_lower:
                # 如果存在但大小写不同，使用实际的列名替换配置中的列名
                actual_column = sub_columns_lower[header.lower()]
                if actual_column != header:
                    logging.info(f"替代料表表头大小写不同，使用实际列名: '{actual_column}' 替代 '{header}'")
                    sub_header_mapping[field] = actual_column
            else:
                # 只检查必需的替代料表字段
                if field in ['pn', 'attribute']:  # 只检查物料编号和属性字段
                    missing_sub_fields['required'].append(header)
                    tkinter.messagebox.showwarning('警告', f'替代料表中未找到必需的表头 "{header}"，请检查表头配置')
                else:  # 其他字段为可选
                    missing_sub_fields['optional'].append(header)
                    tkinter.messagebox.showwarning('警告', f'替代料表中未找到可选的表头 "{header}"，部分信息可能无法显示')

        if missing_sub_fields['required'] or missing_sub_fields['optional']:
            logging.warning(f"替代料表缺少字段: 必需={missing_sub_fields['required']}, 可选={missing_sub_fields['optional']}")

            # 如果缺少Description列，添加一个空列以避免后续处理错误
            if sub_header_mapping['description'] not in sub_df.columns:
                sub_df[sub_header_mapping['description']] = ""
                logging.info(f"已添加空的Description列到替代料表: {sub_header_mapping['description']}")

        # 先对原始BOM的item进行顺序编号
        logging.info("开始对原始BOM进行item重新编号")
        update_status('正在对原始BOM进行item重新编号...')

        # 创建临时列，提取主序号和子序号
        item_col = bom_header_mapping['item']
        bom_df['主序号'] = bom_df[item_col].apply(
            lambda x: int(str(x).split('.')[0]) if not pd.isna(x) and '.' in str(x) else
                     int(x) if not pd.isna(x) and str(x).isdigit() else 999999
        )

        bom_df['子序号'] = bom_df[item_col].apply(
            lambda x: int(str(x).split('.')[1]) if not pd.isna(x) and '.' in str(x) else 0
        )

        # 按主序号和子序号排序
        bom_df = bom_df.sort_values(['主序号', '子序号'])

        # 重新编号（从1开始的连续数字）
        new_items = []
        current_item = 1

        for idx, row in bom_df.iterrows():
            if row['子序号'] == 0:  # 普通行
                new_items.append(str(current_item))
                current_item += 1
            else:  # 已有子序号的行也转为普通序号
                new_items.append(str(current_item))
                current_item += 1

        # 更新Item列
        bom_df[item_col] = new_items

        # 删除临时列
        bom_df = bom_df.drop(['主序号', '子序号'], axis=1)

        # 更新进度
        update_progress(20)

        # TODO: 实现核心处理逻辑

        # 设置默认输出路径
        output_path = Path(bom_path).parent / (Path(bom_path).stem + '_替代料.xlsx')

        # 更新进度（解析完成）
        update_progress(30)

        # 获取映射后的列名
        pn_col = bom_header_mapping['pn']
        ref_col = bom_header_mapping['reference']
        desc_col = bom_header_mapping['description']
        mfr_pn_col = bom_header_mapping['mfr_pn']
        mfr_col = bom_header_mapping['manufacturer']
        quantity_col = bom_header_mapping['quantity']
        part_col = bom_header_mapping['part']

        # 替代料表列名
        sub_pn_col = sub_header_mapping['pn']
        sub_part_col = sub_header_mapping['part']  # 添加零件字段
        sub_desc_col = sub_header_mapping['description']
        sub_mfr_pn_col = sub_header_mapping['mfr_pn']
        sub_mfr_col = sub_header_mapping['manufacturer']
        attr_col = sub_header_mapping['attribute']

        # 检查必需字段
        if pn_col not in bom_df.columns:
            error_msg = f"BOM文件缺少必需列：{pn_col}"
            logging.error(error_msg)
            tkinter.messagebox.showerror('错误', error_msg)
            return

        # 检查替代料表必需字段
        missing_cols = []
        if sub_pn_col not in sub_df.columns:
            missing_cols.append(sub_pn_col)
        if attr_col not in sub_df.columns:
            missing_cols.append(attr_col)

        if missing_cols:
            error_msg = f"替代料表缺少必需列：{', '.join(missing_cols)}"
            logging.error(error_msg)
            tkinter.messagebox.showerror('错误', error_msg)
            return

        # 记录当前使用的字段映射
        logging.info(f"BOM 表头映射: {bom_header_mapping}")
        logging.info(f"替代料表 表头映射: {sub_header_mapping}")
        logging.info(f"替代料表列: {list(sub_df.columns)}")

        # 确定要合并的替代料表字段
        sub_columns_to_merge = [sub_pn_col, attr_col]  # 必需字段
        columns_mapping = {sub_pn_col: pn_col, attr_col: 'sub_attr'}

        # 添加可选字段，如果它们存在于替代料表中
        if sub_part_col in sub_df.columns:
            sub_columns_to_merge.append(sub_part_col)
            columns_mapping[sub_part_col] = part_col
        if sub_desc_col in sub_df.columns:
            sub_columns_to_merge.append(sub_desc_col)
            columns_mapping[sub_desc_col] = desc_col
        if sub_mfr_pn_col in sub_df.columns:
            sub_columns_to_merge.append(sub_mfr_pn_col)
            columns_mapping[sub_mfr_pn_col] = mfr_pn_col
        if sub_mfr_col in sub_df.columns:
            sub_columns_to_merge.append(sub_mfr_col)
            columns_mapping[sub_mfr_col] = mfr_col

        # 合并BOM表和替代料表
        try:
            # 确保列名有效且存在
            sub_columns_filtered = [col for col in sub_columns_to_merge if col in sub_df.columns]
            if sub_columns_filtered:
                # 只使用有效列进行合并
                merged = pd.merge(
                    bom_df,
                    sub_df[sub_columns_filtered].rename(columns={k: v for k, v in columns_mapping.items() if k in sub_columns_filtered}),
                    on=pn_col,
                    how='left'
                )
            else:
                merged = bom_df.copy()
                logging.warning("没有可合并的替代料列，使用原始BOM数据")
        except Exception as e:
            error_msg = translate_error_to_chinese(e)
            logging.error(f"合并BOM和替代料表失败: {e}")

            # 使用自定义错误对话框
            error_details = f"合并BOM和替代料表时出现问题：\n\n{error_msg}\n\n程序将使用原始BOM数据继续处理。"
            show_custom_error('数据处理警告', error_details)
            merged = bom_df.copy()  # 出错时使用原始BOM数据

        # 更新进度（合并完成）
        update_progress(60)

        # 初始化替代料分组变量
        valid_groups = []
        valid_subs = pd.DataFrame()

        # 替代料分组处理 - 只有当替代料表中有必需字段时才执行
        if sub_pn_col in sub_df.columns and attr_col in sub_df.columns:
            logging.info(f"开始替代料分组处理，使用属性字段: {attr_col}")

            try:
                # 筛选有效替代料（相同attribute值）
                # 根据替代料表的attribute值进行分组
                sub_groups = sub_df.groupby(attr_col)

                # 优化：直接使用列表推导获取有效替代组
                valid_groups = [group for _, group in sub_groups if len(group) > 1]

                # 如果有有效替代组，合并为一个DataFrame，避免空DataFrame的警告
                if valid_groups:
                    valid_subs = pd.concat(valid_groups).dropna(axis=1, how='all')
                else:
                    valid_subs = pd.DataFrame()

                logging.info(f"找到 {len(valid_groups)} 个有效替代组")
            except Exception as e:
                error_msg = translate_error_to_chinese(e)
                logging.error(f"处理替代料分组时出错: {e}")

                # 使用自定义错误对话框
                error_details = f"处理替代料分组时出错：\n\n{error_msg}\n\n程序将不应用替代料分组功能。"
                show_custom_error('数据处理警告', error_details)
                valid_groups = []
                valid_subs = pd.DataFrame()
        else:
            # 缺少必需字段时发出警告
            missing_fields = []
            if sub_pn_col not in sub_df.columns:
                missing_fields.append(sub_pn_col)
            if attr_col not in sub_df.columns:
                missing_fields.append(attr_col)

            logging.warning(f"替代料表缺少必需字段: {', '.join(missing_fields)}，跳过替代料分组处理")
            tkinter.messagebox.showwarning('警告', f'替代料表缺少必需字段: {", ".join(missing_fields)}，无法进行替代料处理')

        # 初始化统计变量
        total_count = 0
        matched_count = 0
        unmatched_count = 0

        # 新增统计变量
        original_ref_count = 0  # 原始物料总位号数
        substitute_count = 0    # 替代料的数量

        # 生成新Item序号（原始行+替代行）
        new_items = []
        for idx, row in bom_df.iterrows():
            total_count += 1
            # 计算原始物料的位号数
            if not pd.isna(row[ref_col]) and str(row[ref_col]).strip():
                original_ref_count += count_references(row[ref_col])

            # 查找所有包含当前物料的替代组
            matched_groups = [
                group for group in valid_groups
                if row[pn_col] in group[pn_col].values
            ]

            # 仅在存在替代组时修改原始行Item编号
            if matched_groups:
                new_row = row.copy()
                # 确保 row[item_col] 是字符串类型，并处理可能的NaN值
                if pd.isna(row[item_col]):
                    original_item = "0"  # 如果Item为NaN，使用默认值0
                else:
                    # 尝试拆分Item值
                    try:
                        original_item = str(row[item_col]).split('.')[0]
                    except:
                        original_item = str(row[item_col])  # 如果拆分失败，使用完整值

                # 计算Reference中的位号数量
                reference_text = str(row[ref_col]) if not pd.isna(row[ref_col]) else ''
                ref_count = count_references(reference_text)

                new_row[item_col] = f"{original_item}.1"
                new_row['操作类型'] = '保留'  # 显式设置操作类型
                new_row[ref_col] = reference_text
                new_row[quantity_col] = ref_count  # 根据位号数量更新Quantity
                new_items.append(new_row)

                substitute_counter = 2

                for group in matched_groups:
                    for sub_idx, sub_row in group.iterrows():
                        if sub_row[pn_col] != row[pn_col]:
                            # 计算Reference中的位号数量
                            reference_text = row[ref_col] if not pd.isna(row[ref_col]) else ''
                            ref_count = count_references(reference_text)

                            # 统计替代料数量
                            substitute_count += 1

                            # 创建具有必要字段的替代料项
                            sub_dict = {
                                item_col: f"{original_item}.{substitute_counter}",
                                pn_col: sub_row[pn_col],
                                part_col: sub_row[sub_part_col] if sub_part_col in sub_row.index else row[part_col],  # 优先使用替代料表中的零件字段
                                ref_col: reference_text,
                                quantity_col: ref_count,  # 基于位号数量设置Quantity
                                '操作类型': '替代插入'
                            }

                            # 安全地添加可选字段
                            # 描述字段
                            if desc_col in sub_row.index:
                                sub_dict[desc_col] = sub_row[desc_col]
                            elif desc_col in row.index:
                                sub_dict[desc_col] = row[desc_col]
                            else:
                                sub_dict[desc_col] = ""

                            # 制造商料号字段
                            if mfr_pn_col in sub_row.index:
                                sub_dict[mfr_pn_col] = sub_row[mfr_pn_col]
                            elif mfr_pn_col in row.index:
                                sub_dict[mfr_pn_col] = row[mfr_pn_col]

                            # 制造商字段
                            if mfr_col in sub_row.index:
                                sub_dict[mfr_col] = sub_row[mfr_col]
                            elif mfr_col in row.index:
                                sub_dict[mfr_col] = row[mfr_col]

                            # 如果替代料表中存在属性列，添加到替代料项中
                            if attr_col in sub_row.index:
                                sub_dict[attr_col] = sub_row[attr_col]

                            sub_item = pd.Series(sub_dict)
                            new_items.append(sub_item)
                            substitute_counter += 1
            else:
                # 无替代组时保留原始Item
                new_row = row.copy()
                new_row['操作类型'] = ''  # 无替代组，操作类型为空

                # 计算Reference中的位号数量
                reference_text = str(row[ref_col]) if not pd.isna(row[ref_col]) else ''
                ref_count = count_references(reference_text)

                # 更新Reference和Quantity
                new_row[ref_col] = reference_text
                new_row[quantity_col] = ref_count  # 根据位号数量更新Quantity

                new_items.append(new_row)
                unmatched_count += 1

            # 更新匹配计数（如果有任何替代组）
            if matched_groups:
                matched_count += 1

        # 创建最终DataFrame
        processed_df = pd.DataFrame(new_items)

        # 更新完成进度
        update_progress(90)

        # 计算处理后的总物料数（用于统计）
        total_final_items = len(processed_df)

        # 合并相同P/N的行
        update_status('正在合并相同料号...')

        # 获取P/N列中有重复的值
        duplicate_pns = processed_df[pn_col][processed_df[pn_col].duplicated(keep=False)].unique()

        # 用于存储合并后的数据
        merged_rows = []
        processed_indices = []

        # 跟踪合并物料的详细信息
        merged_materials = []

        # 处理每个重复的P/N
        for pn in duplicate_pns:
            # 获取具有相同P/N的所有行
            duplicate_rows = processed_df[processed_df[pn_col] == pn]

            if len(duplicate_rows) <= 1:
                continue

            # 创建合并后的行（基于第一行）
            merged_row = duplicate_rows.iloc[0].copy()

            # 收集所有Reference，维持原顺序
            all_references = []

            for _, row in duplicate_rows.iterrows():
                processed_indices.append(row.name)

                # 处理Reference，保持原顺序
                if not pd.isna(row[ref_col]) and str(row[ref_col]).strip():
                    refs = [ref.strip() for ref in str(row[ref_col]).split(',')]
                    for ref in refs:
                        if ref and ref not in all_references:  # 只添加非空且不重复的引用
                            all_references.append(ref)

            # 合并后的位号字符串
            combined_references = ','.join(all_references)

            # 设置合并后的值 - Quantity等于位号的数量
            merged_row[ref_col] = combined_references
            # 使用统一的函数计算位号数量
            merged_row[quantity_col] = count_references(combined_references)

            # 记录合并信息
            merge_info = {
                pn_col: pn,
                '合并行数': len(duplicate_rows),
                '合并后位号数': count_references(combined_references)
            }

            # 安全地添加可选字段
            if desc_col in merged_row:
                merge_info[desc_col] = merged_row.get(desc_col, '')
            else:
                merge_info[desc_col] = ''

            if mfr_col in merged_row:
                merge_info[mfr_col] = merged_row.get(mfr_col, '')

            if mfr_pn_col in merged_row:
                merge_info[mfr_pn_col] = merged_row.get(mfr_pn_col, '')

            merged_materials.append(merge_info)

            # 打印调试信息
            logging.info(f"合并料号 {pn}, 合并后位号数量: {count_references(combined_references)}, 位号: {combined_references}")

            merged_rows.append(merged_row)

        # 从原始DataFrame中删除已处理的行
        processed_df = processed_df.drop(processed_indices)

        # 添加合并后的行
        if merged_rows:
            merged_df = pd.DataFrame(merged_rows)
            processed_df = pd.concat([processed_df, merged_df], ignore_index=True)

        # 确保所有行的Quantity都基于Reference位号计数
        processed_df['Quantity_new'] = processed_df.apply(
            lambda row: count_references(row[ref_col]),
            axis=1
        )

        # 用新计算的Quantity替换原Quantity
        processed_df[quantity_col] = processed_df['Quantity_new']
        processed_df.drop('Quantity_new', axis=1, inplace=True)

        # 更新完成进度
        update_progress(95)

        # 按原始Item排序并重新编号
        try:
            logging.info("开始Item排序和重新编号")
            update_status('正在排序和重新编号...')

            # 创建临时列，分别提取主序号和子序号
            processed_df['主序号'] = processed_df[item_col].apply(
                lambda x: int(str(x).split('.')[0]) if not pd.isna(x) and '.' in str(x) else
                         int(x) if not pd.isna(x) and str(x).isdigit() else 999999
            )

            processed_df['子序号'] = processed_df[item_col].apply(
                lambda x: int(str(x).split('.')[1]) if not pd.isna(x) and '.' in str(x) else 0
            )

            # 先按主序号排序，再按子序号排序
            processed_df = processed_df.sort_values(['主序号', '子序号'])

            # 重置索引，便于按顺序处理
            processed_df = processed_df.reset_index(drop=True)

            # 获取所有唯一的主序号，并按顺序排列
            unique_main_numbers = sorted(processed_df['主序号'].unique())

            # 创建主序号映射（老序号->新序号）
            main_number_map = {old: new+1 for new, old in enumerate(unique_main_numbers)}

            # 创建新的DataFrame存储结果
            result_df = pd.DataFrame(columns=processed_df.columns)

            # 新的连续序号
            new_seq = 1

            # 用于记录已经编号的物料组
            processed_sub_groups = set()

            # 处理每个主序号组
            for main_num in unique_main_numbers:
                # 获取当前主序号的所有行
                main_group = processed_df[processed_df['主序号'] == main_num].copy()

                # 新的主序号
                new_main = main_number_map[main_num]

                # 检查组内是否有替代料关系
                # 如果存在'操作类型'列且有值为'替代插入'或'保留'的行，则认为是替代料关系
                has_substitute = '操作类型' in main_group.columns and any(
                    op_type in ['替代插入', '保留'] for op_type in main_group['操作类型'] if not pd.isna(op_type)
                )

                # 检查是否都是同一物料的不同位号
                same_material = False
                if len(main_group) > 1:
                    # 检查所有行是否有相同的P/N
                    unique_pns = main_group[pn_col].unique()
                    same_material = len(unique_pns) == 1

                # 分离带子序号的行和不带子序号的行
                sub_rows = main_group[main_group['子序号'] > 0]
                regular_rows = main_group[main_group['子序号'] == 0]

                # 处理带有替代料标记的行
                if has_substitute and not sub_rows.empty:
                    # 只对真正的替代料关系使用x.1, x.2格式
                    sub_rows_result = sub_rows.copy()
                    for idx, row in sub_rows.iterrows():
                        current_main = row['主序号']
                        current_sub = row['子序号']
                        group_id = f"{current_main}.{current_sub}"
                        if group_id in processed_sub_groups:
                            continue
                        processed_sub_groups.add(group_id)

                        sub_rows_result.loc[idx, item_col] = f"{new_seq}.{current_sub}"

                    # 优化concat操作，避免FutureWarning
                    if not sub_rows_result.empty:
                        sub_rows_filtered = sub_rows_result.dropna(axis=1, how='all')
                        result_df = pd.concat([result_df, sub_rows_filtered])
                    new_seq += 1

                    # 单独处理没有子序号的行，即使主序号相同
                    if not regular_rows.empty:
                        # 优化：批量创建DataFrame而不是逐行添加，减少连接操作次数
                        regular_rows_copies = []
                        for idx, row in regular_rows.iterrows():
                            row_copy = row.copy()
                            row_copy[item_col] = str(new_seq)
                            regular_rows_copies.append(row_copy)
                            new_seq += 1
                        if regular_rows_copies:
                            regular_df = pd.DataFrame(regular_rows_copies)
                            result_df = pd.concat([result_df, regular_df.dropna(axis=1, how='all')])

                # 不是替代料且是同样物料的不同位号 - 仍然按单独序号处理
                elif same_material and main_group['子序号'].sum() == 0:
                    # 相同物料的不同位号，每行单独分配序号
                    # 优化：批量处理
                    material_rows_copies = []
                    for idx, row in main_group.iterrows():
                        row_copy = row.copy()
                        row_copy[item_col] = str(new_seq)
                        material_rows_copies.append(row_copy)
                        new_seq += 1
                    if material_rows_copies:
                        material_df = pd.DataFrame(material_rows_copies)
                        result_df = pd.concat([result_df, material_df.dropna(axis=1, how='all')])

                # 单行无子序号
                elif len(main_group) == 1 and main_group.iloc[0]['子序号'] == 0:
                    main_group[item_col] = str(new_seq)
                    # 优化concat操作，避免FutureWarning
                    if not main_group.empty:
                        # 过滤掉全为NA的列
                        main_group_filtered = main_group.dropna(axis=1, how='all')
                        result_df = pd.concat([result_df, main_group_filtered])
                    new_seq += 1

                # 其他情况（有多行但没有替代料标记）
                else:
                    # 其他情况的多行，每行单独分配序号
                    # 优化：批量处理
                    other_rows_copies = []
                    for idx, row in main_group.iterrows():
                        row_copy = row.copy()
                        row_copy[item_col] = str(new_seq)
                        other_rows_copies.append(row_copy)
                        new_seq += 1
                    if other_rows_copies:
                        other_df = pd.DataFrame(other_rows_copies)
                        result_df = pd.concat([result_df, other_df.dropna(axis=1, how='all')])

            # 使用处理后的DataFrame
            processed_df = result_df.drop(['主序号', '子序号'], axis=1)

            # 最后按Item排序确保顺序正确
            try:
                def natural_sort_key(s):
                    if pd.isna(s):
                        return [0, 0]
                    parts = str(s).split('.')
                    return [int(parts[0]) if parts[0].isdigit() else 0,
                           int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0]

                processed_df = processed_df.sort_values(item_col, key=lambda x: x.map(natural_sort_key))
            except Exception as e:
                logging.warning(f"最终排序失败: {e}，保持当前顺序")

        except Exception as e:
            logging.warning(f"重新编号过程中出现错误: {e}，使用备选排序方法")
            try:
                # 备选排序方法
                def natural_sort_key(s):
                    if pd.isna(s):
                        return [0, 0]
                    parts = str(s).split('.')
                    return [int(parts[0]) if parts[0].isdigit() else 0,
                           int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0]

                processed_df = processed_df.sort_values(item_col, key=lambda x: x.map(natural_sort_key))
            except Exception as e2:
                logging.warning(f"备选排序也失败: {e2}，使用基本排序")
                try:
                    processed_df = processed_df.sort_values(item_col)
                except:
                    logging.warning("所有排序方法均失败，保持原有顺序")

        # 过滤掉DataFrame中的空白列
        logging.info("开始过滤空白列")
        update_status('正在过滤空白列...')

        # 移除所有列都为空的列
        processed_df = processed_df.dropna(axis=1, how='all')

        # 移除不包含任何数据的列（全为空值或者空字符串）
        empty_cols = []
        for col in processed_df.columns:
            # 检查是否所有值都是空值或空字符串
            if processed_df[col].isnull().all() or (processed_df[col].astype(str).str.strip() == '').all():
                empty_cols.append(col)

        # 检查无名列或列名为空格的列
        for col in processed_df.columns:
            if col is None or (isinstance(col, str) and col.strip() == ''):
                empty_cols.append(col)

        # 删除空列
        if empty_cols:
            processed_df = processed_df.drop(columns=empty_cols)
            logging.info(f"已移除 {len(empty_cols)} 个空白列")

        # 确保没有重复的列名
        processed_df = processed_df.loc[:, ~processed_df.columns.duplicated()]
        logging.info("已移除重复列")

        # 检查操作类型列是否有有效数据，如果全为空则删除
        if '操作类型' in processed_df.columns:
            if processed_df['操作类型'].isnull().all() or (processed_df['操作类型'].astype(str).str.strip() == '').all():
                processed_df = processed_df.drop(columns=['操作类型'])
                logging.info("移除无数据的操作类型列")

        # 保存结果
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 写入数据，不包含索引
            processed_df.to_excel(writer, index=False, startrow=len(project_info_rows))

            # 获取工作表
            worksheet = writer.sheets['Sheet1']

            # 获取实际数据列数
            actual_column_count = len(processed_df.columns)

            # 恢复项目信息行
            for row_idx, row_data in enumerate(project_info_rows, 1):
                for col_idx, cell_data in row_data.items():
                    # 只处理实际数据列范围内的单元格
                    if col_idx <= actual_column_count:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = cell_data['value']

                        # 恢复字体
                        cell.font = Font(
                            name=cell_data['font_name'],
                            size=cell_data['font_size'],
                            bold=cell_data['font_bold']
                        )

                        # 恢复填充
                        if cell_data['fill_type'] and cell_data['fill_color']:
                            start_color = cell_data['fill_color']
                            end_color = cell_data['fill_color']
                            new_cell = PatternFill(
                            fill_type=cell_data['fill_type'],
                                start_color=start_color,
                                end_color=end_color
                        )
                            cell.fill = new_cell

                        # 恢复边框
                        border_styles = {
                            'left': cell_data['border_left'],
                            'right': cell_data['border_right'],
                            'top': cell_data['border_top'],
                            'bottom': cell_data['border_bottom']
                        }
                        cell.border = Border(**{
                            side: Side(style=style) if style else None
                            for side, style in border_styles.items()
                        })

                        # 恢复对齐
                        cell.alignment = Alignment(
                            horizontal=cell_data['alignment_horizontal'],
                            vertical=cell_data['alignment_vertical']
                        )

                        # 恢复数字格式
                        cell.number_format = cell_data['number_format']

            # 设置列宽 - 根据表头映射设置
            column_info = {
                'item': {'width': 6, 'index': None},  # Item
                'pn': {'width': 12, 'index': None},  # P/N
                'part': {'width': 12, 'index': None},  # Part
                'reference': {'width': 45, 'index': None},  # Reference
                'quantity': {'width': 10, 'index': None},  # Quantity
                'description': {'width': 50, 'index': None},  # Description
                'mfr_pn': {'width': 22, 'index': None},  # ManuFacturer P/N
                'manufacturer': {'width': 15, 'index': None}  # ManuFacturer
            }

            # 获取每个列的索引位置
            for i, col_name in enumerate(processed_df.columns):
                for key, header in bom_header_mapping.items():
                    if col_name == header and key in column_info:
                        # 列索引从1开始
                        column_info[key]['index'] = i + 1

            # 设置列宽
            for key, info in column_info.items():
                if info['index'] is not None:
                    col_letter = openpyxl.utils.get_column_letter(info['index'])
                    worksheet.column_dimensions[col_letter].width = info['width']

            # 定义样式
            title_font = Font(name='Calibri', size=11, bold=True)
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='Calibri', size=11)
            substitute_font = Font(name='Calibri', size=11, italic=True)

            # 表头样式
            header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')  # 微软蓝

            # 从配置中获取高亮颜色
            highlight_color = config.get('highlight_color', 'FFFF00')  # 默认黄色
            substitute_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type='solid')

            # 边框样式
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

            header_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='005499')  # 底部边框使用深蓝色
            )

            # 对齐样式
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')
            wrap_alignment = Alignment(horizontal='left', vertical='center')  # 移除wrap_text=True

            # 获取操作类型列索引
            op_type_col = processed_df.columns.get_loc('操作类型') + 1 if '操作类型' in processed_df.columns else -1

            # 表头行
            header_row = len(project_info_rows) + 1

            # 应用表头样式（只处理实际数据列）
            for col in range(1, actual_column_count + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = center_alignment

                # 设置特定列的对齐方式
                if col == column_info.get('item', {}).get('index'):
                    cell.alignment = center_alignment  # Item列居中
                elif col == column_info.get('quantity', {}).get('index'):
                    cell.alignment = right_alignment  # 数量列靠右
                elif col == column_info.get('reference', {}).get('index'):
                    cell.alignment = wrap_alignment  # 位号列自动换行
                else:
                    cell.alignment = left_alignment  # 其他列靠左

                # 替代料行的特殊样式
                row_type = worksheet.cell(row=header_row, column=op_type_col).value if op_type_col > 0 else ''
                if row_type == '替代插入':
                    cell.fill = substitute_fill
                    cell.font = substitute_font
                else:
                    cell.font = data_font

            # 应用数据行样式（只处理实际数据列）
            for row in range(header_row + 1, worksheet.max_row + 1):
                row_type = worksheet.cell(row=row, column=op_type_col).value if op_type_col > 0 else ''

                for col in range(1, actual_column_count + 1):
                    cell = worksheet.cell(row=row, column=col)

                    # 设置基本样式
                    cell.border = thin_border

                    # 特定列的对齐方式
                    if col == column_info.get('item', {}).get('index'):
                        cell.alignment = center_alignment  # Item列居中
                    elif col == column_info.get('quantity', {}).get('index'):
                        cell.alignment = right_alignment  # 数量列靠右
                    elif col == column_info.get('reference', {}).get('index'):
                        cell.alignment = wrap_alignment  # 位号列自动换行
                    elif col == column_info.get('description', {}).get('index'):
                        cell.alignment = wrap_alignment  # 描述列自动换行
                    else:
                        cell.alignment = left_alignment  # 其他列靠左

                    # 替代料行的特殊样式
                    if row_type == '替代插入':
                        cell.fill = substitute_fill
                        cell.font = substitute_font
                    else:
                        cell.font = data_font

            # 设置行高
            for row in range(header_row, worksheet.max_row + 1):
                if row == header_row:
                    worksheet.row_dimensions[row].height = 20  # 表头行稍高
                else:
                    worksheet.row_dimensions[row].height = 18  # 数据行统一高度

            # 设置冻结窗格（冻结表头行）
            worksheet.freeze_panes = f'A{header_row + 1}'

            # 添加自动筛选
            # ref_cell = f'A{header_row}:{openpyxl.utils.get_column_letter(worksheet.max_column)}{header_row}'
            # worksheet.auto_filter.ref = ref_cell

            # 设置工作表标题
            worksheet.title = "BOM"

            # 复制原始BOM文件中的其他工作表（包含样式）
            logging.info("开始复制原始BOM文件中的其他工作表（包含样式）")
            try:
                # 打开原始BOM文件
                original_wb = openpyxl.load_workbook(bom_path)

                # 遍历所有工作表
                for sheet_name in original_wb.sheetnames:
                    # 跳过主工作表（已处理）
                    if sheet_name == original_wb.active.title:
                        continue

                    logging.info(f"复制工作表: {sheet_name}")

                    # 复制工作表到新文件
                    if sheet_name not in writer.book.sheetnames:
                        # 获取原始工作表
                        source_sheet = original_wb[sheet_name]

                        # 创建新工作表
                        target_sheet = writer.book.create_sheet(title=sheet_name)

                        # 复制单元格数据和样式
                        for row_idx, row in enumerate(source_sheet.rows, 1):
                            for col_idx, source_cell in enumerate(row, 1):
                                # 创建新单元格并复制值
                                target_cell = target_sheet.cell(row=row_idx, column=col_idx, value=source_cell.value)

                                # 复制字体
                                if source_cell.font:
                                    target_cell.font = Font(
                                        name=source_cell.font.name,
                                        size=source_cell.font.size,
                                        bold=source_cell.font.bold,
                                        italic=source_cell.font.italic,
                                        underline=source_cell.font.underline,
                                        strike=source_cell.font.strike,
                                        color=source_cell.font.color
                                    )

                                # 复制对齐方式
                                if source_cell.alignment:
                                    target_cell.alignment = Alignment(
                                        horizontal=source_cell.alignment.horizontal,
                                        vertical=source_cell.alignment.vertical,
                                        textRotation=source_cell.alignment.textRotation,
                                        wrapText=source_cell.alignment.wrapText,
                                        shrinkToFit=source_cell.alignment.shrinkToFit,
                                        indent=source_cell.alignment.indent
                                    )

                                # 复制边框
                                if source_cell.border:
                                    sides = {}
                                    for side in ['left', 'right', 'top', 'bottom']:
                                        side_obj = getattr(source_cell.border, side)
                                        if side_obj and side_obj.style:
                                            sides[side] = Side(style=side_obj.style, color=side_obj.color)
                                        else:
                                            sides[side] = None

                                    target_cell.border = Border(**sides)

                                # 复制填充
                                if source_cell.fill and source_cell.fill.fill_type != 'none':
                                    try:
                                        fill_type = source_cell.fill.fill_type

                                        # 创建新的填充对象
                                        if fill_type == 'solid' or fill_type == 'solid':
                                            if hasattr(source_cell.fill, 'start_color') and source_cell.fill.start_color:
                                                rgb = source_cell.fill.start_color.rgb if hasattr(source_cell.fill.start_color, 'rgb') else None
                                                if rgb:
                                                    target_cell.fill = PatternFill(fill_type='solid', start_color=rgb)
                                    except Exception as fill_error:
                                        logging.warning(f"复制填充样式失败: {fill_error}")

                                # 复制数字格式
                                if source_cell.number_format:
                                    target_cell.number_format = source_cell.number_format

                        # 复制工作表级别的属性

                        # 复制列宽
                        for col_letter, column_dimensions in source_sheet.column_dimensions.items():
                            if column_dimensions.width is not None:
                                target_sheet.column_dimensions[col_letter].width = column_dimensions.width

                                # 复制列的hidden属性
                                if hasattr(column_dimensions, 'hidden'):
                                    target_sheet.column_dimensions[col_letter].hidden = column_dimensions.hidden

                        # 复制行高和行的隐藏状态
                        for row_num, row_dimensions in source_sheet.row_dimensions.items():
                            if row_dimensions.height is not None:
                                target_sheet.row_dimensions[row_num].height = row_dimensions.height

                            # 复制行的hidden属性
                            if hasattr(row_dimensions, 'hidden'):
                                target_sheet.row_dimensions[row_num].hidden = row_dimensions.hidden

                        # 复制合并单元格
                        for merged_range in source_sheet.merged_cells.ranges:
                            target_sheet.merge_cells(str(merged_range))

                        # 复制打印设置
                        if hasattr(source_sheet, 'page_setup') and hasattr(target_sheet, 'page_setup'):
                            target_sheet.page_setup.orientation = source_sheet.page_setup.orientation
                            target_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
                            target_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
                            target_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth

                        # 复制视图设置
                        if hasattr(source_sheet, 'sheet_view') and hasattr(target_sheet, 'sheet_view'):
                            target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
                            target_sheet.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale

                        # 复制冻结窗格设置
                        if source_sheet.freeze_panes:
                            target_sheet.freeze_panes = source_sheet.freeze_panes

                        logging.info(f"已复制工作表(含样式): {sheet_name}")
            except Exception as e:
                logging.error(f"复制工作表时出错: {e}", exc_info=True)
                logging.info("尝试使用备用方法复制工作表（仅数据）")
                try:
                    # 备用方法：只复制数据
                    if sheet_name not in writer.book.sheetnames:
                        # 获取原始工作表
                        source_sheet = original_wb[sheet_name]

                        # 创建新工作表
                        target_sheet = writer.book.create_sheet(title=sheet_name)

                        # 只复制单元格数据和基本属性
                        for row in source_sheet.rows:
                            for cell in row:
                                target_sheet.cell(row=cell.row, column=cell.column).value = cell.value

                        # 复制列宽
                        for col_letter, column_dimensions in source_sheet.column_dimensions.items():
                            if column_dimensions.width is not None:
                                target_sheet.column_dimensions[col_letter].width = column_dimensions.width

                        # 复制行高
                        for row_num, row_dimensions in source_sheet.row_dimensions.items():
                            if row_dimensions.height is not None:
                                target_sheet.row_dimensions[row_num].height = row_dimensions.height

                        logging.info(f"已复制工作表(仅数据): {sheet_name}")
                except Exception as backup_error:
                    logging.error(f"备用复制方法也失败: {backup_error}", exc_info=True)

        # 更新进度为100%完成
        update_progress(100)

        logging.info(f'处理完成，输出文件已保存至：{output_path}')

        # 计算处理时间
        end_time = time.time()
        process_duration = end_time - start_time
        # 格式化时间显示
        if process_duration < 60:
            time_str = f"{process_duration:.2f}秒"
        else:
            minutes = int(process_duration // 60)
            seconds = process_duration % 60
            time_str = f"{minutes}分{seconds:.2f}秒"

        # 美化统计信息显示
        stats_info = []

        # ===== 主标题 =====
        stats_info.append("✅ 处理完成！")
        stats_info.append("-" * 40)

        # ===== 基本统计信息 =====
        stats_info.append("📊 基本统计")
        stats_info.append(f"• 总物料数: {total_count}个")
        stats_info.append(f"• 匹配替代料: {matched_count}个")
        stats_info.append(f"• 未匹配物料: {unmatched_count}个")
        stats_info.append(f"• 处理时长: {time_str}")
        stats_info.append(f"• 输出文件: {output_path}")

        # ===== 替代料统计 =====
        stats_info.append("\n📋 替代料统计")
        stats_info.append("-" * 40)
        stats_info.append(f"• 添加替代料数量: {substitute_count}个")
        stats_info.append(f"• 添加替代料后总物料数: {total_final_items}个")
        stats_info.append(f"• 原始物料总位号数: {original_ref_count}个")

        # 计算处理后的总位号数（不含替代料）
        final_ref_count = sum(count_references(str(row[ref_col])) for _, row in processed_df.iterrows() if row.get('操作类型', '') != '替代插入')
        stats_info.append(f"• 处理后物料总位号数: {final_ref_count}个")

        # ===== 物料合并信息 =====
        if merged_materials:
            stats_info.append("\n🔄 相同物料合并信息")
            stats_info.append("-" * 40)

            # 添加合并汇总信息
            total_merged_rows = sum(mat['合并行数'] for mat in merged_materials)
            total_merged_refs = sum(mat['合并后位号数'] for mat in merged_materials)
            stats_info.append(f"• 共合并{len(merged_materials)}种物料，{total_merged_rows}行 → {len(merged_materials)}行")
            stats_info.append(f"• 合并后总位号数: {total_merged_refs}个")

            # 显示所有合并物料的详细信息
            if merged_materials:
                stats_info.append("\n详细合并信息:")

            for idx, mat in enumerate(merged_materials, 1):
                    stats_info.append(f"\n  物料 {idx}:")
                    stats_info.append(f"  • {pn_col}: {str(mat[pn_col])}")
                    if mat[desc_col]:
                        # 裁剪描述文本，避免过长
                        desc = str(mat[desc_col])  # 确保desc是字符串类型
                        if len(desc) > 50:
                            desc = desc[:47] + "..."
                        stats_info.append(f"  • 描述: {desc}")
                    if mat[mfr_pn_col]:
                        stats_info.append(f"  • 制造商料号: {str(mat[mfr_pn_col])}")
                    if mat[mfr_col]:
                        stats_info.append(f"  • 制造商: {str(mat[mfr_col])}")
                    stats_info.append(f"  • 合并: {str(mat['合并行数'])}行 → {str(mat['合并后位号数'])}个位号")

        # 合并成格式化的文本
        formatted_stats = "\n".join(stats_info)

        # 更新状态文本
        update_status(formatted_stats)

    except Exception as e:
        error_msg = translate_error_to_chinese(e)
        logging.error(f'处理失败：{str(e)}', exc_info=True)
        update_progress(0)
        update_status(f'处理失败：{error_msg}')

        # 使用自定义错误对话框显示错误
        error_details = f"错误类型：{type(e).__name__}\n\n错误描述：{error_msg}\n\n如果问题仍然存在，请联系开发者获取支持。"
        show_custom_error('处理失败', error_details)

        # 不再抛出异常，避免程序崩溃
        return

def reset_default_sub_path():
    """重置默认替代料表路径"""
    config = load_config()
    if config['default_sub_path']:
        if tkinter.messagebox.askyesno('重置默认路径', '是否重置默认替代料表路径？'):
            config['default_sub_path'] = ''
            save_config(config)
            tkinter.messagebox.showinfo('重置成功', '已重置默认替代料表路径。')
    else:
        tkinter.messagebox.showinfo('提示', '当前未设置默认替代料表路径。')

def reset_all_config():
    """重置所有配置"""
    global _config_cache

    # 使用内置默认配置，不从外部文件加载
    default_config = load_default_config(use_builtin_defaults=True)

    # 确保只保留需要的配置项
    default_config = {
        'last_bom_dir': '',
        'default_sub_path': '',
        'bom_header_mapping': default_config['bom_header_mapping'],
        'sub_header_mapping': default_config['sub_header_mapping'],
        'highlight_color': default_config['highlight_color'],
        'last_update_check': 0,  # 重置上次检查更新的时间戳
        'last_used_header_mapping': {}  # 重置上次使用的表头映射
    }

    try:
        # 更新缓存
        _config_cache = default_config

        # 使用save_config函数保存配置，它会同时保存到用户配置文件和程序目录下的config.json文件
        save_config(default_config)

        # 显示成功消息
        logging.info("已重置所有配置为默认值")
        tkinter.messagebox.showinfo("重置所有配置", "已重置所有配置到程序内置默认值")
        return True
    except Exception as e:
        logging.error(f"重置配置失败: {e}")
        tkinter.messagebox.showerror('错误', f'重置配置失败: {e}')
        return False

def show_header_config():
    """显示表头配置对话框"""
    config = load_config()
    bom_header_mapping = config.get('bom_header_mapping', {})
    sub_header_mapping = config.get('sub_header_mapping', {})

    # 打印当前配置，用于调试
    print("显示表头配置对话框时的配置:")
    print("BOM表头映射:", bom_header_mapping)
    print("替代料表表头映射:", sub_header_mapping)
    print("高亮颜色:", config.get('highlight_color', 'FFFF00'))

    # 创建配置窗口
    config_window = tk.Toplevel(root)
    config_window.title("表头配置")
    config_window.geometry("600x550")  # 增加窗口高度以容纳颜色选择

    # 设置窗口居中
    window_width = 600
    window_height = 550  # 更新窗口高度
    screen_width = config_window.winfo_screenwidth()
    screen_height = config_window.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    config_window.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

    config_window.grab_set()  # 模态窗口

    # 主框架
    main_frame = ttk.Frame(config_window, padding=20)
    main_frame.pack(fill='both', expand=True)

    # 创建选项卡控件
    tab_control = ttk.Notebook(main_frame)
    tab_control.pack(fill='both', expand=True, pady=(0, 20))

    # 创建BOM表头配置选项卡
    bom_tab = ttk.Frame(tab_control, padding=10)
    tab_control.add(bom_tab, text=" BOM文件表头 ")

    # 创建替代料表表头配置选项卡
    sub_tab = ttk.Frame(tab_control, padding=10)
    tab_control.add(sub_tab, text=" 替代料表表头 ")

    # 创建颜色设置选项卡
    color_tab = ttk.Frame(tab_control, padding=10)
    tab_control.add(color_tab, text=" 颜色设置 ")

    # === BOM表头配置 ===
    ttk.Label(bom_tab, text="配置BOM文件各字段的表头名称",
              font=('微软雅黑', 10, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 15))

    # BOM表头描述和输入框
    bom_field_descriptions = {
        'item': 'Item编号字段',
        'pn': '物料编号字段',
        'part': '零件字段',
        'reference': '位号字段',
        'quantity': '数量字段',
        'description': '描述字段',
        'mfr_pn': '制造商料号字段',
        'manufacturer': '制造商字段'
    }

    # BOM表头输入框
    bom_header_entries = {}

    for i, (key, description) in enumerate(bom_field_descriptions.items()):
        ttk.Label(bom_tab, text=description + ":",
                 anchor='e').grid(row=i+1, column=0, sticky='e', padx=(0, 10), pady=5)

        entry = ttk.Entry(bom_tab, width=30)
        entry.grid(row=i+1, column=1, sticky='w', pady=5)
        entry.insert(0, bom_header_mapping.get(key, ''))
        bom_header_entries[key] = entry

    # === 替代料表表头配置 ===
    ttk.Label(sub_tab, text="配置替代料表各字段的表头名称",
              font=('微软雅黑', 10, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 15))

    # 强调提示
    ttk.Label(sub_tab, text="注意: '替代料属性字段' 仅用于替代料表，不需要在BOM文件中存在",
              font=('微软雅黑', 9), foreground='red').grid(row=1, column=0, columnspan=2, sticky='w', pady=(0, 10))

    # 替代料表表头描述和输入框
    sub_field_descriptions = {
        'pn': '物料编号字段',
        'part': '零件字段',
        'description': '描述字段',
        'mfr_pn': '制造商料号字段',
        'manufacturer': '制造商字段',
        'attribute': '替代料属性字段'
    }

    # 替代料表表头输入框
    sub_header_entries = {}

    for i, (key, description) in enumerate(sub_field_descriptions.items()):
        row_idx = i + 2  # 因为添加了一行提示，所以从第2行开始
        ttk.Label(sub_tab, text=description + ":",
                 anchor='e').grid(row=row_idx, column=0, sticky='e', padx=(0, 10), pady=5)

        entry = ttk.Entry(sub_tab, width=30)
        entry.grid(row=row_idx, column=1, sticky='w', pady=5)
        entry.insert(0, sub_header_mapping.get(key, ''))
        sub_header_entries[key] = entry

    # === 颜色设置 ===
    ttk.Label(color_tab, text="配置替代料高亮颜色",
              font=('微软雅黑', 10, 'bold')).grid(row=0, column=0, columnspan=3, sticky='w', pady=(0, 15))

    # 颜色选择变量
    color_var = tk.StringVar(value=config.get('highlight_color', 'FFFF00'))

    # 颜色选项
    colors = [
        ('黄色', 'FFFF00'),
        ('浅黄色', 'FFFFC0'),
        ('浅绿色', 'C6EFCE'),
        ('浅蓝色', 'BDD7EE'),
        ('浅红色', 'FFB6C1'),
        ('浅灰色', 'D9D9D9')
    ]

    # 创建颜色预览和选择按钮，分两行显示
    for i, (color_name, color_code) in enumerate(colors):
        # 确定行和列位置（每行3个）
        row_pos = 1 + i // 3
        col_pos = i % 3

        # 创建颜色选择框架
        color_frame = ttk.Frame(color_tab)
        color_frame.grid(row=row_pos, column=col_pos, padx=10, pady=5, sticky='w')

        # 颜色预览框
        preview = tk.Label(color_frame, width=4, height=2)
        preview.configure(bg=f'#{color_code}')
        preview.pack(side='top', pady=(0, 3))

        # 单选按钮
        rb = ttk.Radiobutton(color_frame, text=color_name,
                           variable=color_var, value=color_code)
        rb.pack(side='top')

    # 按钮框架
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(side='bottom', pady=10)

    # 保存按钮
    save_button = ttk.Button(
        button_frame,
        text="保存配置",
        command=lambda: save_header_config(bom_header_entries, sub_header_entries, config_window, color_var.get())
    )
    save_button.pack(side='left', padx=5)

    # 恢复默认按钮
    reset_button = ttk.Button(
        button_frame,
        text="恢复默认",
        command=lambda: reset_header_config(bom_header_entries, sub_header_entries, color_var, config_window)
    )
    reset_button.pack(side='left', padx=5)

    # 重置所有配置按钮
    reset_all_button = ttk.Button(
        button_frame,
        text="重置所有配置",
        command=lambda: [config_window.destroy(), reset_all_config()]
    )
    reset_all_button.pack(side='left', padx=5)

    # 取消按钮
    cancel_button = ttk.Button(
        button_frame,
        text="取消",
        command=config_window.destroy
    )
    cancel_button.pack(side='left', padx=5)

def save_header_config(bom_entries, sub_entries, window, highlight_color):
    """保存表头配置"""
    config = load_config()

    # 获取BOM表头配置
    for key, entry in bom_entries.items():
        value = entry.get().strip()
        if value:  # 只更新非空值
            config['bom_header_mapping'][key] = value

    # 获取替代料表表头配置
    for key, entry in sub_entries.items():
        value = entry.get().strip()
        if value:  # 只更新非空值
            config['sub_header_mapping'][key] = value

    # 保存高亮颜色设置
    config['highlight_color'] = highlight_color

    # 使用save_config函数保存配置，它会同时保存到用户配置文件和程序目录下的config.json文件
    save_config(config)

    # 显示成功消息
    tkinter.messagebox.showinfo("保存成功", "配置已保存")

    # 关闭窗口
    window.destroy()

def reset_header_config(bom_entries, sub_entries, color_var, window=None):
    """
    重置表头配置为默认值并自动保存

    Args:
        bom_entries: BOM表头输入框字典
        sub_entries: 替代料表表头输入框字典
        color_var: 颜色变量
        window: 配置窗口，如果提供则在保存后关闭窗口
    """
    # 使用内置默认配置，不从外部文件加载
    default_config = load_default_config(use_builtin_defaults=True)

    # 获取默认表头映射
    default_bom_header_mapping = default_config['bom_header_mapping']
    default_sub_header_mapping = default_config['sub_header_mapping']

    # 将默认值填入BOM表头输入框
    for key, entry in bom_entries.items():
        entry.delete(0, tk.END)
        entry.insert(0, default_bom_header_mapping.get(key, ''))

    # 将默认值填入替代料表表头输入框
    for key, entry in sub_entries.items():
        entry.delete(0, tk.END)
        entry.insert(0, default_sub_header_mapping.get(key, ''))

    # 使用内置默认颜色
    color_var.set(default_config.get('highlight_color', 'FFFF00'))

    # 自动保存配置
    config = load_config()

    # 更新配置
    for key, entry in bom_entries.items():
        config['bom_header_mapping'][key] = entry.get()

    for key, entry in sub_entries.items():
        config['sub_header_mapping'][key] = entry.get()

    config['highlight_color'] = color_var.get()

    # 保存配置
    save_config(config)

    # 显示成功消息
    tkinter.messagebox.showinfo("恢复默认", "已恢复到程序内置默认配置并自动保存")

    # 如果提供了窗口参数，关闭窗口
    if window:
        window.destroy()



# 更新检测相关函数
def check_for_updates(current_version):
    """
    检查GitHub上是否有新版本

    Args:
        current_version: 当前版本号

    Returns:
        tuple: (是否有更新, 最新版本, 下载链接, 更新日志, 是否为exe更新)
    """
    try:
        print(f"检查更新，当前版本: {current_version}")

        # 设置请求头，避免API限制
        headers = {
            "User-Agent": "BOM-Tool-Update-Checker"
        }

        # 添加超时设置，避免长时间等待
        response = requests.get(GITHUB_API_URL, headers=headers, timeout=DOWNLOAD_TIMEOUT)

        if response.status_code == 200:
            data = response.json()
            latest_version = data["tag_name"].lstrip("v")
            print(f"发现版本: {latest_version}")

            # 使用packaging.version进行版本比较
            if pkg_version.parse(latest_version) > pkg_version.parse(current_version):
                print(f"发现新版本: {latest_version}")

                # 查找exe资源文件
                download_url = ""
                is_exe_update = False

                for asset in data.get("assets", []):
                    if asset["name"].endswith(".exe"):
                        download_url = asset["browser_download_url"]
                        is_exe_update = True
                        print(f"找到exe更新: {asset['name']}")
                        break

                # 如果没有资源文件，使用源代码下载链接
                if not download_url:
                    download_url = data["zipball_url"]
                    print("使用源代码链接作为备用")

                # 获取更新日志
                changelog = data["body"] if "body" in data else "无可用的更新日志"

                return True, latest_version, download_url, changelog, is_exe_update

        # 如果没有新版本或请求失败
        return False, current_version, "", "", False
    except Exception as e:
        error_msg = translate_error_to_chinese(e)
        print(f"检查更新失败: {str(e)}")
        logging.error(f"检查更新失败: {str(e)}")
        return False, current_version, "", f"检查更新失败: {error_msg}", False

def download_with_resume(url, dest_file, progress_callback=None, status_callback=None):
    """
    支持断点续传的下载函数

    Args:
        url: 下载链接
        dest_file: 目标文件路径
        progress_callback: 进度回调函数，接收三个参数(已下载大小, 总大小, 进度百分比)
        status_callback: 状态回调函数，接收一个参数(状态消息)

    Returns:
        bool: 下载是否成功
    """
    # 检查是否存在部分下载的文件
    file_size = 0
    if os.path.exists(dest_file):
        file_size = os.path.getsize(dest_file)
        if status_callback:
            status_callback(f"发现已下载的文件({file_size/1024:.1f}KB)，继续下载...")

    # 设置HTTP头，支持断点续传
    headers = {}
    if file_size > 0:
        headers['Range'] = f'bytes={file_size}-'

    # 打开文件，使用追加模式
    with open(dest_file, 'ab' if file_size > 0 else 'wb') as f:
        retries = 0
        while retries < DOWNLOAD_MAX_RETRIES:
            try:
                # 发起请求
                response = requests.get(url, headers=headers, stream=True, timeout=DOWNLOAD_TIMEOUT)

                # 检查响应状态码
                if file_size > 0 and response.status_code == 416:
                    # 范围请求错误，文件可能已经完整下载
                    if status_callback:
                        status_callback("文件已完整下载")
                    return True
                elif file_size > 0 and response.status_code != 206:
                    # 不支持断点续传，重新下载
                    if status_callback:
                        status_callback("服务器不支持断点续传，重新下载...")
                    f.close()
                    os.remove(dest_file)
                    return download_with_resume(url, dest_file, progress_callback, status_callback)
                elif response.status_code not in [200, 206]:
                    # 其他错误
                    raise Exception(f"下载失败，HTTP状态码: {response.status_code}")

                # 获取文件总大小
                total_size = int(response.headers.get('content-length', 0)) + file_size
                if total_size == 0:
                    total_size = file_size  # 如果无法获取总大小，使用已下载大小

                # 已下载大小
                downloaded = file_size

                # 下载文件
                for chunk in response.iter_content(chunk_size=DOWNLOAD_CHUNK_SIZE):
                    if chunk:  # 过滤掉保持连接的空块
                        f.write(chunk)
                        downloaded += len(chunk)

                        # 计算进度
                        progress = 0
                        if total_size > 0:
                            progress = int(downloaded * 100 / total_size)

                        # 更新进度
                        if progress_callback:
                            progress_callback(downloaded, total_size, progress)

                # 下载完成
                if status_callback:
                    status_callback("下载完成")
                return True

            except (requests.exceptions.RequestException, IOError) as e:
                retries += 1
                if status_callback:
                    status_callback(f"下载出错，正在重试 ({retries}/{DOWNLOAD_MAX_RETRIES}): {str(e)}")

                # 如果不是最后一次重试，等待一段时间再重试
                if retries < DOWNLOAD_MAX_RETRIES:
                    time.sleep(2 * retries)  # 指数退避

        # 超过最大重试次数
        if status_callback:
            status_callback("下载失败，超过最大重试次数")
        return False

def show_update_notification(parent, current_version, latest_version, changelog, download_url, is_exe_update):
    """
    显示更新通知对话框

    Args:
        parent: 父窗口
        current_version: 当前版本
        latest_version: 最新版本
        changelog: 更新日志
        download_url: 下载链接
        is_exe_update: 是否为exe更新

    Returns:
        bool: 用户是否选择更新
    """
    # 创建更新通知对话框
    dialog = tk.Toplevel(parent)
    dialog.title("发现新版本")
    dialog.geometry("500x400")
    dialog.transient(parent)  # 设置为父窗口的子窗口
    dialog.grab_set()  # 模态对话框

    # 设置窗口居中
    window_width = 500
    window_height = 400
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    dialog.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

    # 主框架
    main_frame = ttk.Frame(dialog, padding=20)
    main_frame.pack(fill='both', expand=True)

    # 版本信息
    version_frame = ttk.Frame(main_frame)
    version_frame.pack(fill='x', pady=(0, 15))

    ttk.Label(version_frame, text=f"当前版本: v{current_version}",
              font=('微软雅黑', 10)).pack(side='left')

    ttk.Label(version_frame, text=f"最新版本: v{latest_version}",
              font=('微软雅黑', 10, 'bold'), foreground='#0078D4').pack(side='right')

    # 分隔线
    ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=5)

    # 更新内容
    ttk.Label(main_frame, text="更新内容:",
              font=('微软雅黑', 10, 'bold')).pack(anchor='w', pady=(10, 5))

    # 更新日志文本框
    changelog_frame = ttk.Frame(main_frame)
    changelog_frame.pack(fill='both', expand=True, pady=(0, 15))

    # 滚动条
    scrollbar = ttk.Scrollbar(changelog_frame)
    scrollbar.pack(side='right', fill='y')

    # 文本框
    changelog_text = tk.Text(changelog_frame, wrap=tk.WORD, height=10,
                           yscrollcommand=scrollbar.set,
                           font=('微软雅黑', 9), background='#F9F9F9')
    changelog_text.pack(fill='both', expand=True)
    scrollbar.config(command=changelog_text.yview)

    # 插入更新日志
    changelog_text.insert(tk.END, changelog)
    changelog_text.config(state=tk.DISABLED)  # 设置为只读

    # 下载信息
    download_frame = ttk.Frame(main_frame)
    download_frame.pack(fill='x', pady=(0, 15))

    if is_exe_update:
        download_text = "可执行文件更新，下载后将自动安装"
    else:
        download_text = "源代码更新，下载后需手动安装"

    ttk.Label(download_frame, text=download_text,
              font=('微软雅黑', 9), foreground='#666666').pack(anchor='w')

    # 按钮框架
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill='x', pady=(0, 10))

    # 用户选择结果
    result = [False]  # 使用列表存储结果，以便在回调函数中修改

    # 更新按钮
    def on_update():
        result[0] = True
        dialog.destroy()

    update_button = ttk.Button(button_frame, text="立即更新", command=on_update, width=15)
    update_button.pack(side='left', padx=(0, 10))

    # 取消按钮
    cancel_button = ttk.Button(button_frame, text="稍后再说", command=dialog.destroy, width=15)
    cancel_button.pack(side='left')

    # 等待用户操作
    parent.wait_window(dialog)

    return result[0]

class UpdateManager:
    """
    更新管理器类，负责检查更新、下载更新和安装更新
    """
    def __init__(self, master):
        self.master = master
        self.version = APP_VERSION

        # 更新状态
        self.update_available = False
        self.latest_version = ""
        self.download_url = ""
        self.update_changelog = ""
        self.is_exe_update = False

        # 更新窗口状态
        self.update_window_open = False

        # 文本颜色
        self.text_color = "#000000"

    def check_updates_on_startup(self):
        """程序启动时检查更新"""
        # 延迟几秒，让主界面先加载完成
        time.sleep(2)

        # 检查上次更新时间，如果距离上次检查时间不足UPDATE_CHECK_INTERVAL天，则不检查
        config = load_config()
        last_check = config.get('last_update_check', 0)
        now = time.time()

        if now - last_check < UPDATE_CHECK_INTERVAL * 24 * 60 * 60:
            return

        # 检查更新
        has_update, latest_version, download_url, changelog, is_exe_update = check_for_updates(self.version)

        # 更新最后检查时间
        config['last_update_check'] = now
        save_config(config)

        if has_update:
            # 保存更新信息
            self.update_available = True
            self.latest_version = latest_version
            self.download_url = download_url
            self.update_changelog = changelog
            self.is_exe_update = is_exe_update

            # 显示更新提示
            self.master.after(0, self.show_update_notification)

    def check_updates_manually(self):
        """手动检查更新"""
        # 更新状态栏
        self._update_status("正在检查更新...", "#0078D4")

        # 在新线程中检查更新
        Thread(target=self._check_updates_thread, args=(True,)).start()

    def _check_updates_thread(self, is_manual_check=False):
        """检查更新的线程函数"""
        try:
            # 检查更新
            has_update, latest_version, download_url, changelog, is_exe_update = check_for_updates(self.version)

            # 如果是手动检查，更新状态栏
            if is_manual_check:
                self._update_status("检查更新完成", self.text_color)

            # 如果有更新，显示更新通知
            if has_update:
                # 保存更新信息
                self.update_available = True
                self.latest_version = latest_version
                self.download_url = download_url
                self.update_changelog = changelog
                self.is_exe_update = is_exe_update

                # 在主线程中显示更新通知
                self.master.after(0, self.show_update_notification)
            elif is_manual_check:
                # 如果是手动检查且没有更新，显示提示
                self.master.after(0, lambda: messagebox.showinfo("检查更新",
                                                            f"当前版本 {self.version} 已是最新版本。"))
        except Exception as e:
            error_msg = translate_error_to_chinese(e)
            print(f"检查更新时出错: {str(e)}")
            logging.error(f"检查更新时出错: {str(e)}")
            if is_manual_check:
                self._update_status("检查更新失败", "#FF0000")
                # 使用自定义错误对话框
                error_details = f"检查更新时出错：\n\n{error_msg}\n\n请检查网络连接或稍后重试。"
                self.master.after(0, lambda: show_custom_error("检查更新失败", error_details, self.master))

    def show_update_notification(self):
        """显示更新通知"""
        # 如果更新窗口已经打开，则返回
        if self.update_window_open:
            return

        if messagebox.askyesno("发现新版本",
                            f"发现新版本 v{self.latest_version}，当前版本 v{self.version}。\n\n是否查看更新内容并更新？"):
            self.show_update_dialog()

    def show_update_dialog(self):
        """显示更新对话框"""
        # 如果更新窗口已经打开，则返回
        if self.update_window_open:
            return

        # 设置窗口打开标志
        self.update_window_open = True

        # 显示更新通知对话框
        if show_update_notification(self.master, self.version, self.latest_version,
                                  self.update_changelog, self.download_url, self.is_exe_update):
            # 用户选择更新，开始下载
            self._download_update(self.latest_version, self.download_url, self.is_exe_update)
        else:
            # 用户取消更新，重置窗口打开标志
            self.update_window_open = False

    def _download_update(self, latest_version, download_url, is_exe_update):
        """下载更新"""
        # 确定下载路径
        if is_exe_update:
            # 如果是可执行文件，下载到临时目录
            download_dir = tempfile.gettempdir()
            # 从URL中提取文件名
            file_name = os.path.basename(download_url)
            if not file_name.endswith('.exe'):
                # 如果URL中没有文件名，使用默认文件名
                file_name = f"BOM替代料工具_v{latest_version}.exe"
            download_path = os.path.join(download_dir, file_name)
        else:
            # 如果是源代码，下载到用户选择的目录
            download_dir = filedialog.askdirectory(title="选择保存目录")
            if not download_dir:
                # 用户取消选择，取消下载
                self.update_window_open = False
                return
            download_path = os.path.join(download_dir, f"BOM替代料工具_v{latest_version}.zip")

        # 创建进度对话框
        progress_dialog = tk.Toplevel(self.master)
        progress_dialog.title("下载更新")
        progress_dialog.geometry("400x150")
        progress_dialog.transient(self.master)  # 设置为主窗口的子窗口
        progress_dialog.grab_set()  # 模态对话框

        # 设置窗口居中
        window_width = 400
        window_height = 150
        screen_width = progress_dialog.winfo_screenwidth()
        screen_height = progress_dialog.winfo_screenheight()
        center_x = int((screen_width - window_width) / 2)
        center_y = int((screen_height - window_height) / 2)
        progress_dialog.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

        # 主框架
        main_frame = ttk.Frame(progress_dialog, padding=20)
        main_frame.pack(fill='both', expand=True)

        # 下载信息
        info_label = ttk.Label(main_frame, text=f"正在下载 v{latest_version}...",
                              font=('微软雅黑', 10))
        info_label.pack(anchor='w', pady=(0, 10))

        # 进度条
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(main_frame, orient='horizontal',
                                     length=360, mode='determinate',
                                     variable=progress_var)
        progress_bar.pack(fill='x', pady=(0, 5))

        # 进度百分比
        progress_label = ttk.Label(main_frame, text="0%", font=('微软雅黑', 9))
        progress_label.pack(anchor='e')

        # 状态信息
        status_label = ttk.Label(main_frame, text="准备下载...",
                               font=('微软雅黑', 9), foreground='#666666')
        status_label.pack(anchor='w', pady=(5, 0))

        # 进度回调函数
        def update_progress_callback(downloaded, total, progress):
            progress_dialog.after(0, lambda: [
                progress_var.set(progress),
                progress_label.config(text=f"{progress}%"),
                info_label.config(text=f"正在下载 v{latest_version}... {downloaded/1024/1024:.1f}MB/{total/1024/1024:.1f}MB")
            ])

        # 状态回调函数
        def status_callback(message):
            progress_dialog.after(0, lambda: status_label.config(text=message))

        # 在新线程中下载
        def download_thread():
            try:
                # 下载文件
                success = download_with_resume(download_url, download_path,
                                             update_progress_callback, status_callback)

                # 如果下载成功
                if success:
                    # 关闭进度对话框
                    progress_dialog.destroy()

                    # 显示下载完成对话框
                    if is_exe_update:
                        # 如果是可执行文件，询问用户是否关闭当前程序并运行新版本
                        if messagebox.askyesno("更新完成",
                                            f"新版本 {latest_version} 已下载完成。\n\n"
                                            f"是否关闭当前程序并运行新版本？"):
                            # 启动新版本并关闭当前程序
                            subprocess.Popen([download_path])
                            self.master.quit()
                            self.master.destroy()
                            sys.exit(0)
                    else:
                        # 如果是源代码包，提示用户下载完成
                        messagebox.showinfo("下载完成",
                                         f"新版本 {latest_version} 已下载到:\n{download_path}")
                else:
                    # 如果下载失败，显示错误消息
                    messagebox.showerror("下载失败",
                                     f"下载新版本 {latest_version} 失败。\n"
                                     f"请稍后重试或访问官方网站手动下载。")
            except Exception as e:
                # 如果发生异常，显示错误消息
                error_msg = translate_error_to_chinese(e)
                logging.error(f"下载过程中发生错误: {str(e)}")

                # 使用自定义错误对话框
                error_details = f"下载更新文件时出错：\n\n{error_msg}\n\n请检查网络连接或稍后重试。您也可以手动下载更新文件。"
                show_custom_error("下载错误", error_details)
            finally:
                # 重置窗口打开标志
                self.update_window_open = False

        # 启动下载线程
        Thread(target=download_thread).start()

    def _update_status(self, message, color=None):
        """更新状态栏"""
        # 这个方法需要在主应用程序中实现
        # 在这里只是一个占位符
        pass

# 修改主程序入口
if __name__ == '__main__':
    setup_logging()

    # 打印系统信息，帮助诊断
    print("\n=== 系统信息 ===")
    print(f"Python版本: {sys.version}")
    print(f"操作系统: {platform.platform()}")
    print(f"系统架构: {platform.architecture()}")
    print(f"当前工作目录: {os.getcwd()}")

    # 检查是否是打包环境
    is_frozen = getattr(sys, 'frozen', False)
    print(f"是否是打包环境: {is_frozen}")
    if is_frozen:
        print(f"可执行文件路径: {sys.executable}")
        print(f"可执行文件目录: {os.path.dirname(sys.executable)}")

    # 尝试获取用户目录信息
    try:
        print(f"用户主目录: {os.path.expanduser('~')}")
        print(f"用户文档目录: {os.path.join(os.path.expanduser('~'), 'Documents')}")
    except Exception as e:
        print(f"获取用户目录信息失败: {e}")

    print("=== 系统信息结束 ===\n")

    # 检查启动参数
    reset_config = False

    # 检查是否有命令行参数
    if len(sys.argv) > 1:
        print(f"命令行参数: {sys.argv[1:]}")
        if sys.argv[1] == '--reset-config' or sys.argv[1] == '-r':
            reset_config = True

    # 检查是否同时按下Shift键
    import ctypes
    reset_by_key = False
    try:
        # 检查Shift键状态
        shift_state = ctypes.windll.user32.GetAsyncKeyState(0x10) & 0x8000 != 0
        if shift_state:
            print("检测到Shift键被按下")
            logging.info("检测到Shift键被按下")
            reset_by_key = True
    except Exception as e:
        print(f"检查Shift键状态失败: {e}")
        logging.error(f"检查Shift键状态失败: {e}")

    # 重置配置
    if reset_config or reset_by_key:
        # 删除配置文件，完全重置
        try:
            if os.path.exists(CONFIG_FILE):
                print(f"正在删除配置文件: {CONFIG_FILE}")
                os.remove(CONFIG_FILE)
                msg = "配置文件已删除，将使用默认配置"
                print(msg)
                logging.info(msg)
                # 如果是按键触发的，显示消息框
                if reset_by_key:
                    tkinter.messagebox.showinfo('重置成功', msg)
            else:
                msg = f"未找到配置文件: {CONFIG_FILE}，将使用默认配置"
                print(msg)
                logging.info(msg)

            # 创建默认配置并保存
            default_config = get_builtin_default_config()
            save_result = save_config(default_config)
            print(f"创建默认配置文件结果: {save_result}")
            logging.info(f"创建默认配置文件结果: {save_result}")
        except Exception as e:
            error_msg = f"重置配置失败: {e}"
            print(error_msg)
            logging.error(error_msg)

    # 检查程序目录是否可写
    print("\n=== 配置文件目录检查 ===")

    # 检查程序目录
    program_dir = get_program_dir()
    print(f"程序目录: {program_dir}")
    program_dir_writable = check_directory_writable(program_dir)
    print(f"程序目录可写: {program_dir_writable}")
    logging.info(f"程序目录可写: {program_dir_writable}")

    # 检查CONFIG_FILE目录
    config_dir = os.path.dirname(CONFIG_FILE)
    print(f"配置文件目录: {config_dir}")
    config_dir_writable = check_directory_writable(config_dir)
    print(f"配置文件目录可写: {config_dir_writable}")
    logging.info(f"配置文件目录可写: {config_dir_writable}")

    if not program_dir_writable:
        print("警告: 程序目录不可写，配置文件将无法保存到程序目录！")
        logging.warning("程序目录不可写，配置文件将无法保存到程序目录！")

        # 检查当前工作目录
        current_dir = os.getcwd()
        if program_dir != current_dir:
            current_dir_writable = check_directory_writable(current_dir)
            print(f"当前工作目录: {current_dir}")
            print(f"当前工作目录可写: {current_dir_writable}")
            logging.info(f"当前工作目录可写: {current_dir_writable}")

            if current_dir_writable:
                print(f"将尝试保存配置到当前工作目录: {current_dir}")
                logging.info(f"将尝试保存配置到当前工作目录: {current_dir}")
            else:
                print("警告: 当前工作目录也不可写，配置文件将无法保存！")
                logging.warning("当前工作目录也不可写，配置文件将无法保存！")

                # 尝试用户目录
                try:
                    user_home = os.path.expanduser("~")
                    user_home_writable = check_directory_writable(user_home)
                    print(f"用户主目录: {user_home}")
                    print(f"用户主目录可写: {user_home_writable}")
                    logging.info(f"用户主目录可写: {user_home_writable}")

                    if user_home_writable:
                        print(f"将尝试保存配置到用户主目录: {user_home}")
                        logging.info(f"将尝试保存配置到用户主目录: {user_home}")
                except Exception as e:
                    print(f"检查用户主目录失败: {e}")
                    logging.error(f"检查用户主目录失败: {e}")

    print("=== 配置文件目录检查完成 ===\n")

    # 创建GUI
    create_gui()

    # 启动时检查更新（在新线程中运行）
    Thread(target=update_manager.check_updates_on_startup).start()

    # 处理完成时播放提示音
    import winsound
    winsound.MessageBeep()

    # 错误时播放不同提示音
    # winsound.MessageBeep(winsound.MB_ICONHAND)